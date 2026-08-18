#!/usr/bin/env python3
"""Build district panels + simplified Wahlkreis GeoJSON for BE / ST (/ MV geo helper).

Panels match the MV swing CSV schema used by code/district_forecast.py:
  wahlkreis,wahlkreisname,gültige_stimmen_erst,...,gültige_stimmen_zweit,...
"""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

import geopandas as gpd
import openpyxl

REPO = Path(__file__).resolve().parents[1]

PARTIES = ("spd", "afd", "cdu", "linke", "gruene", "fdp", "bsw", "others")

BEZ_NAMES = {
    "01": "Mitte",
    "02": "Friedrichshain-Kreuzberg",
    "03": "Pankow",
    "04": "Charlottenburg-Wilmersdorf",
    "05": "Spandau",
    "06": "Steglitz-Zehlendorf",
    "07": "Tempelhof-Schöneberg",
    "08": "Neukölln",
    "09": "Treptow-Köpenick",
    "10": "Marzahn-Hellersdorf",
    "11": "Lichtenberg",
    "12": "Reinickendorf",
}


def _i(v) -> int:
    if v is None or v == "":
        return 0
    return int(round(float(v)))


def write_panel(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "wahlkreis",
        "wahlkreisname",
        "gültige_stimmen_erst",
        *[f"{p}_erst" for p in PARTIES],
        "gültige_stimmen_zweit",
        *[f"{p}_zweit" for p in PARTIES],
    ]
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for r in sorted(rows, key=lambda x: int(x["wahlkreis"])):
            w.writerow(r)
    print(f"Wrote {path} ({len(rows)} districts)")


def simplify_geo(
    gdf: gpd.GeoDataFrame,
    out_path: Path,
    tolerance: float = 50.0,
) -> None:
    """Project-aware simplify → WGS84 FeatureCollection with wkr / wkr_name."""
    if gdf.crs is None:
        raise ValueError("GeoDataFrame has no CRS")
    metric = gdf.to_crs(3857) if gdf.crs.is_geographic else gdf
    simplified = metric.copy()
    simplified["geometry"] = metric.geometry.simplify(tolerance, preserve_topology=True)
    out = simplified.to_crs(4326)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    # Keep only website props.
    feat = []
    for _, row in out.iterrows():
        feat.append(
            {
                "type": "Feature",
                "properties": {"wkr": int(row["wkr"]), "wkr_name": str(row["wkr_name"])},
                "geometry": json.loads(gpd.GeoSeries([row.geometry]).to_json())["features"][0][
                    "geometry"
                ],
            }
        )
    payload = {"type": "FeatureCollection", "features": feat}
    out_path.write_text(json.dumps(payload, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"Wrote {out_path} ({len(feat)} features, {out_path.stat().st_size // 1024} KB)")


def prepare_st() -> None:
    raw = (REPO / "sachsen-anhalt" / "raw" / "lt21dat1.csv").read_bytes().decode("latin-1")
    reader = csv.DictReader(raw.splitlines(), delimiter=";")
    rows_out = []
    for r in reader:
        if r.get("Satzart") != "WKR":
            continue
        wkr = int(r["Schlüsselnummer"])
        name = r["Name"].strip()
        ge = _i(r["D - Gültige Erststimmen"])
        gz = _i(r["F - Gültige Zweitstimmen"])
        erst = {
            "cdu": _i(r["D01 - CDU "]),
            "afd": _i(r["D02 - AfD "]),
            "linke": _i(r["D03 - DIE LINKE "]),
            "spd": _i(r["D04 - SPD "]),
            "gruene": _i(r["D05 - GRÜNE "]),
            "fdp": _i(r["D06 - FDP "]),
            "bsw": 0,
        }
        zweit = {
            "cdu": _i(r["F01 - CDU "]),
            "afd": _i(r["F02 - AfD "]),
            "linke": _i(r["F03 - DIE LINKE "]),
            "spd": _i(r["F04 - SPD "]),
            "gruene": _i(r["F05 - GRÜNE "]),
            "fdp": _i(r["F06 - FDP "]),
            "bsw": 0,
        }
        erst["others"] = max(0, ge - sum(erst.values()))
        zweit["others"] = max(0, gz - sum(zweit.values()))
        row = {
            "wahlkreis": wkr,
            "wahlkreisname": name,
            "gültige_stimmen_erst": ge,
            "gültige_stimmen_zweit": gz,
        }
        for p in PARTIES:
            row[f"{p}_erst"] = erst[p]
            row[f"{p}_zweit"] = zweit[p]
        rows_out.append(row)
    write_panel(REPO / "sachsen-anhalt" / "ltw21_st_abs.csv", rows_out)

    gdf = gpd.read_file(REPO / "sachsen-anhalt" / "geo" / "raw" / "Wahlkreise_LTW_2026.geojson")
    gdf = gdf.rename(columns={"Nr. Wahlkr": "wkr", "Name Wahlk": "wkr_name"})
    gdf["wkr"] = gdf["wkr"].astype(int)
    # Fix known typo in official geojson.
    gdf.loc[gdf["wkr"] == 27, "wkr_name"] = "Dessau-Roßlau-Wittenberg"
    full = REPO / "sachsen-anhalt" / "geo" / "ltw_wahlkreise_st.geojson"
    simple = REPO / "sachsen-anhalt" / "geo" / "ltw_wahlkreise_st_simple.geojson"
    out = REPO / "output" / "ltw_wahlkreise_st.geojson"
    gdf_wgs = gdf.to_crs(4326)
    full.parent.mkdir(parents=True, exist_ok=True)
    gdf_wgs[["wkr", "wkr_name", "geometry"]].to_file(full, driver="GeoJSON")
    simplify_geo(gdf[["wkr", "wkr_name", "geometry"]], simple, tolerance=80)
    simplify_geo(gdf[["wkr", "wkr_name", "geometry"]], out, tolerance=80)


def prepare_be() -> None:
    """2023 results remapped to 2026 WKs (Zweit); Erst from 2023 where local WK matches."""
    remapped = REPO / "berlin" / "raw" / "be_agh2026_2023.xlsx"
    raw2023 = REPO / "berlin" / "raw" / "be_agh2023.xlsx"

    wb = openpyxl.load_workbook(remapped, read_only=True, data_only=True)
    ws = wb["2023_Zweitstimme"]
    z_tot: dict[tuple[str, int], dict] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row[7] is None or row[1] is None:
            continue
        bez = str(row[1]).zfill(2)
        local = int(row[7])
        key = (bez, local)
        slot = z_tot.setdefault(key, {"valid": 0, **{p: 0 for p in PARTIES}})
        slot["valid"] += _i(row[12])
        slot["cdu"] += _i(row[13])
        slot["spd"] += _i(row[15])
        slot["gruene"] += _i(row[17])
        slot["linke"] += _i(row[19])
        slot["afd"] += _i(row[21])
        slot["fdp"] += _i(row[23])
        # Sonstige column 25 — prefer residual for consistency
    wb.close()
    for slot in z_tot.values():
        known = sum(slot[p] for p in PARTIES if p not in ("others", "bsw"))
        slot["bsw"] = 0
        slot["others"] = max(0, slot["valid"] - known)

    # 2023 Erst by (bez, local) — approximate personal-vote gap on matching local numbers.
    e_tot: dict[tuple[str, int], dict] = {}
    wb = openpyxl.load_workbook(raw2023, read_only=True, data_only=True)
    ws = wb["AGH_W1"]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row[7] is None or row[2] is None:
            continue
        bez = str(row[2]).zfill(2)
        local = int(row[7])
        key = (bez, local)
        slot = e_tot.setdefault(key, {"valid": 0, **{p: 0 for p in PARTIES}})
        slot["valid"] += _i(row[16])  # Gültige Stimmen
        slot["spd"] += _i(row[18])
        slot["cdu"] += _i(row[19])
        slot["gruene"] += _i(row[20])
        slot["linke"] += _i(row[21])
        slot["afd"] += _i(row[22])
        slot["fdp"] += _i(row[23])
        # remaining parties → others (incl. EB)
    wb.close()
    for slot in e_tot.values():
        known = sum(slot[p] for p in PARTIES if p not in ("others", "bsw"))
        slot["bsw"] = 0
        slot["others"] = max(0, slot["valid"] - known)

    # Stable statewide WK numbers: sorted awk codes → 1..78
    awk_keys = sorted(z_tot.keys(), key=lambda x: (x[0], x[1]))
    awk_to_wkr = {k: i + 1 for i, k in enumerate(awk_keys)}

    rows_out = []
    for key in awk_keys:
        bez, local = key
        wkr = awk_to_wkr[key]
        name = f"{BEZ_NAMES.get(bez, bez)} {local}"
        z = z_tot[key]
        if key in e_tot and e_tot[key]["valid"] > 0:
            e = e_tot[key]
        else:
            # New / remapped-only WK: no personal-vote gap.
            e = {p: z[p] for p in PARTIES}
            e["valid"] = z["valid"]
        row = {
            "wahlkreis": wkr,
            "wahlkreisname": name,
            "gültige_stimmen_erst": e["valid"],
            "gültige_stimmen_zweit": z["valid"],
        }
        for p in PARTIES:
            row[f"{p}_erst"] = e[p]
            row[f"{p}_zweit"] = z[p]
        rows_out.append(row)
    write_panel(REPO / "berlin" / "agh23_be_abs.csv", rows_out)

    # Geo from WFS export
    gdf = gpd.read_file(REPO / "berlin" / "geo" / "raw" / "agh2026_awk.geojson")
    # CRS is EPSG:25833
    if gdf.crs is None:
        gdf = gdf.set_crs(25833)

    def awk_key(awk: str) -> tuple[str, int]:
        s = str(awk).zfill(4)
        return s[:2], int(s[2:])

    gdf["awk_key"] = gdf["awk"].map(awk_key)
    missing = [k for k in gdf["awk_key"] if k not in awk_to_wkr]
    if missing:
        raise SystemExit(f"Geo awk keys not in panel: {missing[:5]}")
    gdf["wkr"] = gdf["awk_key"].map(awk_to_wkr)
    gdf["wkr_name"] = gdf["awk_key"].map(
        lambda k: f"{BEZ_NAMES.get(k[0], k[0])} {k[1]}"
    )

    full = REPO / "berlin" / "geo" / "ltw_wahlkreise_be.geojson"
    simple = REPO / "berlin" / "geo" / "ltw_wahlkreise_be_simple.geojson"
    out = REPO / "output" / "ltw_wahlkreise_be.geojson"
    gdf_wgs = gdf.to_crs(4326)
    full.parent.mkdir(parents=True, exist_ok=True)
    gdf_wgs[["wkr", "wkr_name", "geometry"]].to_file(full, driver="GeoJSON")
    simplify_geo(gdf[["wkr", "wkr_name", "geometry"]], simple, tolerance=40)
    simplify_geo(gdf[["wkr", "wkr_name", "geometry"]], out, tolerance=40)

    # Save awk ↔ wkr map for debugging
    meta = {
        "awk_to_wkr": {f"{b}{local:02d}": w for (b, local), w in awk_to_wkr.items()},
        "source_zweit": "DL_BE_AGH2026_AGH2023.xlsx (remapped Zweit)",
        "source_erst": "DL_BE_AGHBVV2023.xlsx AGH_W1 (2023 Erst; gap=0 if WK only in 2026)",
    }
    (REPO / "berlin" / "awk_wkr_map.json").write_text(
        json.dumps(meta, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--state", choices=["BE", "ST", "all"], default="all")
    args = ap.parse_args()
    if args.state in ("ST", "all"):
        prepare_st()
    if args.state in ("BE", "all"):
        prepare_be()


if __name__ == "__main__":
    main()
