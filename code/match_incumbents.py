#!/usr/bin/env python3
"""Identify 2026 candidates who are sitting MPs / prior Direktmandat winners.

Local analysis only — writes under tmp/incumbents/, nothing to website JSON.

Sources
-------
1. Prior Direkt winners (parsed from election-result raw files)
   - BE 2023: berlin/raw/SB_B07-02-03_2023.xlsx  ("Gewählt ist: …")
   - ST 2021: sachsen-anhalt/raw/lt21dat1.csv     (Wahlkreissieger/-in)
   - MV 2021: none in repo (votes-only)

2. Sitting MPs (Wikipedia member lists, current Wahlperiode)
   - BE / MV / ST — name + party + Direkt vs Liste seat

Matching reuses listen_candidates.normalize_name / first_last_key against
2026 Direkt + Listen rosters (placeholders excluded).

Usage
-----
  python code/match_incumbents.py            # fetch wiki + match, write CSVs
  python code/match_incumbents.py --offline  # reuse cached wiki HTML in tmp/
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
import urllib.request
from collections import Counter, defaultdict
from io import StringIO
from pathlib import Path

import openpyxl
import pandas as pd

REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO / "code"))

from listen_candidates import (  # noqa: E402
    STATE_DIRS,
    clean_person_name,
    first_last_key,
    normalize_name,
)

OUT_DIR = REPO / "tmp" / "incumbents"
CACHE_DIR = OUT_DIR / "wiki_cache"

WIKI_URLS = {
    "BE": (
        "https://de.wikipedia.org/wiki/"
        "Liste_der_Mitglieder_des_Abgeordnetenhauses_von_Berlin_(19._Wahlperiode)"
    ),
    "MV": (
        "https://de.wikipedia.org/wiki/"
        "Liste_der_Mitglieder_des_Landtages_Mecklenburg-Vorpommern_(8._Wahlperiode)"
    ),
    "ST": (
        "https://de.wikipedia.org/wiki/"
        "Liste_der_Mitglieder_des_Landtages_von_Sachsen-Anhalt_(8._Wahlperiode)"
    ),
}

PARTY_ALIASES = {
    "grüne": "gruene",
    "grune": "gruene",
    "gruen": "gruene",
    "bündnis 90/die grünen": "gruene",
    "bundnis 90/die grunen": "gruene",
    "bündnis90/die grünen": "gruene",
    "die linke": "linke",
    "linke": "linke",
    "spd": "spd",
    "cdu": "cdu",
    "afd": "afd",
    "fdp": "fdp",
    "gruppe fdp": "fdp",
    "bsw": "bsw",
    "freie wähler": "fw",
    "fw": "fw",
}


def _canon_party(raw: str) -> str:
    s = (raw or "").strip()
    if not s:
        return ""
    # Drop "fraktionslos (...)" → try inner party, else empty
    low = s.lower()
    if low.startswith("fraktionslos"):
        m = re.search(r"\(([^)]+)\)", s)
        if m:
            return _canon_party(m.group(1))
        return ""
    # Strip trailing "-Fraktion" / " Fraktion"
    s = re.sub(r"[\s\-]*[Ff]raktion(?:/Gruppe)?$", "", s).strip()
    key = (
        s.lower()
        .replace("ä", "ae")
        .replace("ö", "oe")
        .replace("ü", "ue")
        .replace("ß", "ss")
    )
    key = re.sub(r"\s+", " ", key).strip()
    return PARTY_ALIASES.get(key, key.replace(" ", "_"))


def _parse_winner_paren(s: str) -> tuple[str, str] | None:
    """'Last, First (PARTY)' → (display_name, party_slug)."""
    m = re.match(r"(.+?)\s*\(([^)]+)\)\s*$", (s or "").strip())
    if not m:
        return None
    return clean_person_name(m.group(1)), _canon_party(m.group(2))


# ---------------------------------------------------------------------------
# Prior Direkt winners
# ---------------------------------------------------------------------------


def extract_be_2023_winners() -> list[dict]:
    path = REPO / "berlin" / "raw" / "SB_B07-02-03_2023.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    pat = re.compile(r"Gewählt ist:\s*(.+?)\s*\(([^)]+)\)")
    out: list[dict] = []
    for sheet in wb.sheetnames:
        if not re.match(r"^3\.\d+$", sheet):
            continue
        wkr = int(sheet.split(".")[1])
        for i, row in enumerate(wb[sheet].iter_rows(values_only=True)):
            if i > 8:
                break
            for v in row:
                if not isinstance(v, str):
                    continue
                m = pat.search(v)
                if not m:
                    continue
                out.append(
                    {
                        "state": "BE",
                        "election": "2023",
                        "wkr": wkr,
                        "name": clean_person_name(m.group(1)),
                        "party": _canon_party(m.group(2)),
                        "seat_type": "direkt",
                        "source": str(path.relative_to(REPO)),
                    }
                )
    return out


def extract_st_2021_winners() -> list[dict]:
    path = REPO / "sachsen-anhalt" / "raw" / "lt21dat1.csv"
    df = pd.read_csv(path, sep=";", encoding="latin-1", dtype=str)
    out: list[dict] = []
    for _, r in df[df["Satzart"] == "WKR"].iterrows():
        parsed = _parse_winner_paren(r["Wahlkreissieger/-in"])
        if not parsed:
            continue
        name, party = parsed
        out.append(
            {
                "state": "ST",
                "election": "2021",
                "wkr": int(r["Schlüsselnummer"]),
                "wkr_name": (r.get("Name") or "").strip(),
                "name": name,
                "party": party,
                "seat_type": "direkt",
                "source": str(path.relative_to(REPO)),
            }
        )
    return out


# ---------------------------------------------------------------------------
# Sitting MPs (Wikipedia)
# ---------------------------------------------------------------------------


def _fetch_wiki(state: str, offline: bool) -> str:
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cache = CACHE_DIR / f"{state.lower()}.html"
    if offline:
        if not cache.exists():
            raise FileNotFoundError(f"--offline but missing {cache}")
        return cache.read_text(encoding="utf-8")
    url = WIKI_URLS[state]
    req = urllib.request.Request(
        url, headers={"User-Agent": "website-pipeline/incumbent-match (research)"}
    )
    html = urllib.request.urlopen(req, timeout=60).read().decode("utf-8")
    cache.write_text(html, encoding="utf-8")
    return html


_WK_NUM_RE = re.compile(r"^(\d+)\b")


def _parse_mv_seat(raw: str) -> tuple[str, int | None]:
    s = (raw or "").strip().replace("\xa0", " ")
    if not s:
        return "unknown", None
    if s.lower().startswith("landesliste"):
        return "liste", None
    m = _WK_NUM_RE.match(s)
    if m:
        return "direkt", int(m.group(1))
    return "unknown", None


def _parse_st_seat(raw: str) -> tuple[str, int | None]:
    s = (raw or "").strip().replace("\xa0", " ")
    if not s or s.lower().startswith("landesliste"):
        return "liste", None
    m = _WK_NUM_RE.match(s)
    if m:
        return "direkt", int(m.group(1))
    return "liste", None


def _parse_be_seat(wahlkreis: str, liste: str) -> tuple[str, str]:
    """Return (seat_type, seat_label). BE wiki lists both Liste and Wahlkreis."""
    wk = (wahlkreis or "").strip().replace("\xa0", " ")
    li = (liste or "").strip().replace("\xa0", " ")
    # If Erststimmen share cell is absent we still have Wahlkreis column for
    # everyone who contested a WK; Wikipedia marks Direktmandat via non-empty
    # Erststimmenanteil in the full table — read_html keeps Wahlkreis for list
    # MPs who also ran in a WK. Prefer: if Liste says Platz and we have no
    # clear signal, treat as liste unless Anmerkungen say otherwise.
    # Practical rule used here: if Liste starts with Landesliste/Bezirksliste
    # AND Wahlkreis looks like a district name, they may still hold a Direktmandat.
    # Wikipedia BE table: Direktmandat winners have Erststimmenanteil; list MPs
    # often have "—" or empty. We don't have that cleaned — use Liste primary.
    if li.lower().startswith("landesliste") or li.lower().startswith("bezirksliste"):
        # Many Direkt winners also appear with a list place (dual candidacy).
        # Without Erststimmen we can't tell — mark as "sitting" only; seat_type
        # "mixed" when both present.
        if wk and not wk.lower().startswith("nan"):
            return "sitting", f"liste={li}; wkr={wk}"
        return "liste", li
    if wk:
        return "direkt", wk
    return "unknown", li or wk


def parse_sitting_mps(state: str, html: str) -> list[dict]:
    tables = pd.read_html(StringIO(html))
    out: list[dict] = []
    source = WIKI_URLS[state]

    if state == "BE":
        # Current members = first large table (excludes Ausgeschiedene).
        t = tables[0]
        for _, r in t.iterrows():
            name = clean_person_name(str(r.get("Name") or ""))
            party = _canon_party(str(r.get("Fraktion") or ""))
            if not name or name.lower() == "nan":
                continue
            seat_type, seat_label = _parse_be_seat(
                str(r.get("Wahlkreis") or ""), str(r.get("Liste") or "")
            )
            out.append(
                {
                    "state": "BE",
                    "name": name,
                    "party": party,
                    "seat_type": seat_type,
                    "seat_label": seat_label,
                    "wkr": "",
                    "source": source,
                }
            )
    elif state == "MV":
        t = tables[2]
        for _, r in t.iterrows():
            name = clean_person_name(str(r.get("Name") or ""))
            party = _canon_party(str(r.get("Fraktion/Gruppe") or ""))
            if not name or name.lower() == "nan":
                continue
            seat_type, wkr = _parse_mv_seat(str(r.get("Wahlkreis bzw. Landesliste") or ""))
            out.append(
                {
                    "state": "MV",
                    "name": name,
                    "party": party,
                    "seat_type": seat_type,
                    "seat_label": str(r.get("Wahlkreis bzw. Landesliste") or "").replace(
                        "\xa0", " "
                    ),
                    "wkr": wkr if wkr is not None else "",
                    "source": source,
                }
            )
    elif state == "ST":
        t = tables[2]
        for _, r in t.iterrows():
            name = clean_person_name(str(r.get("Mitglied des Landtages") or ""))
            party = _canon_party(str(r.get("Fraktion") or ""))
            if not name or name.lower() == "nan":
                continue
            seat_type, wkr = _parse_st_seat(str(r.get("Landtagswahlkreis") or ""))
            out.append(
                {
                    "state": "ST",
                    "name": name,
                    "party": party,
                    "seat_type": seat_type,
                    "seat_label": str(r.get("Landtagswahlkreis") or "").replace(
                        "\xa0", " "
                    ),
                    "wkr": wkr if wkr is not None else "",
                    "source": source,
                }
            )
    return out


# ---------------------------------------------------------------------------
# 2026 candidates + matching
# ---------------------------------------------------------------------------


def load_2026_candidates(state: str) -> list[dict]:
    ddir = STATE_DIRS[state]
    rows: list[dict] = []
    direkt = ddir / "direktkandidaten_2026.csv"
    if direkt.exists():
        with direkt.open(encoding="utf-8", newline="") as f:
            for r in csv.DictReader(f):
                if not (r.get("name") or "").strip():
                    continue
                rows.append(
                    {
                        "cand_type": "direkt",
                        "party": r["party"],
                        "name": clean_person_name(r["name"]),
                        "wkr": r.get("wkr", ""),
                        "list_pos": "",
                        "person_id": "",
                        "source": r.get("source", ""),
                    }
                )
    listen = ddir / "listenkandidaten_2026.csv"
    if listen.exists():
        with listen.open(encoding="utf-8", newline="") as f:
            for r in csv.DictReader(f):
                if r.get("is_placeholder") in ("1", "true", "True"):
                    continue
                if not (r.get("name") or "").strip():
                    continue
                rows.append(
                    {
                        "cand_type": "liste",
                        "party": r["party"],
                        "name": clean_person_name(r["name"]),
                        "wkr": r.get("wkr_direct") or "",
                        "list_pos": r.get("list_pos", ""),
                        "person_id": r.get("person_id", ""),
                        "source": r.get("source", ""),
                    }
                )
    return rows


def _index_candidates(
    cands: list[dict],
) -> tuple[dict[tuple[str, str], list[dict]], dict[tuple[str, tuple[str, str]], list[dict]]]:
    by_norm: dict[tuple[str, str], list[dict]] = defaultdict(list)
    by_fl: dict[tuple[str, tuple[str, str]], list[dict]] = defaultdict(list)
    for c in cands:
        n = normalize_name(c["name"])
        by_norm[(c["party"], n)].append(c)
        fl = first_last_key(c["name"])
        if fl:
            by_fl[(c["party"], fl)].append(c)
    return by_norm, by_fl


def match_people(
    people: list[dict],
    cands: list[dict],
    *,
    role: str,
) -> tuple[list[dict], list[dict]]:
    """Match incumbents/winners to 2026 candidates within the same party."""
    by_norm, by_fl = _index_candidates(cands)
    matched: list[dict] = []
    unmatched: list[dict] = []

    for p in people:
        party = p.get("party") or ""
        if not party:
            # fraktionslos — try name-only unique match across parties
            hits, how = _cross_party_unique(p["name"], cands)
        else:
            hits = by_norm.get((party, normalize_name(p["name"])), [])
            how = "exact"
            if not hits:
                fl = first_last_key(p["name"])
                hits = by_fl.get((party, fl), []) if fl else []
                how = "first_last"

        if not hits:
            unmatched.append({**p, "role": role})
            continue

        # Deduplicate by (cand_type, party, name, wkr, list_pos)
        seen: set[tuple] = set()
        for h in hits:
            key = (h["cand_type"], h["party"], h["name"], h["wkr"], h["list_pos"])
            if key in seen:
                continue
            seen.add(key)
            matched.append(
                {
                    "role": role,
                    "state": p.get("state", ""),
                    "incumbent_name": p["name"],
                    "incumbent_party": party,
                    "incumbent_seat_type": p.get("seat_type", ""),
                    "incumbent_seat_label": p.get("seat_label", ""),
                    "incumbent_wkr": p.get("wkr", ""),
                    "incumbent_election": p.get("election", ""),
                    "match_how": how,
                    "cand_type": h["cand_type"],
                    "cand_party": h["party"],
                    "cand_name": h["name"],
                    "cand_wkr": h["wkr"],
                    "cand_list_pos": h["list_pos"],
                    "cand_person_id": h.get("person_id", ""),
                    "cand_source": h.get("source", ""),
                }
            )
    return matched, unmatched


def _cross_party_unique(
    name: str, cands: list[dict]
) -> tuple[list[dict], str]:
    n = normalize_name(name)
    fl = first_last_key(name)
    exact = [c for c in cands if normalize_name(c["name"]) == n]
    if exact:
        parties = {c["party"] for c in exact}
        if len(parties) == 1:
            return exact, "exact_cross_party"
    if fl:
        fl_hits = [c for c in cands if first_last_key(c["name"]) == fl]
        parties = {c["party"] for c in fl_hits}
        if fl_hits and len(parties) == 1:
            return fl_hits, "first_last_cross_party"
    return [], ""


# ---------------------------------------------------------------------------
# I/O
# ---------------------------------------------------------------------------


def _write_csv(path: Path, rows: list[dict], fieldnames: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    fields = fieldnames or list(rows[0].keys())
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def _person_keys(rows: list[dict], name_key: str, party_key: str) -> set[tuple[str, str]]:
    return {
        (normalize_name(r[name_key]), (r.get(party_key) or ""))
        for r in rows
        if r.get(name_key)
    }


def run(states: list[str], offline: bool) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    winners: list[dict] = []
    if "BE" in states:
        winners.extend(extract_be_2023_winners())
    if "ST" in states:
        winners.extend(extract_st_2021_winners())
    _write_csv(OUT_DIR / "direkt_winners.csv", winners)
    print(f"Direkt winners: {len(winners)} → {OUT_DIR / 'direkt_winners.csv'}")

    all_mps: list[dict] = []
    all_matches: list[dict] = []
    all_unmatched_mps: list[dict] = []
    all_unmatched_winners: list[dict] = []

    summary_rows: list[dict] = []

    for state in states:
        print(f"\n=== {state} ===")
        html = _fetch_wiki(state, offline=offline)
        mps = parse_sitting_mps(state, html)
        all_mps.extend(mps)
        cands = load_2026_candidates(state)
        named = [c for c in cands if c["name"]]
        print(f"  sitting MPs (wiki): {len(mps)}")
        print(f"  2026 named candidates: {len(named)}")

        m_sit, u_sit = match_people(mps, named, role="sitting_mp")
        all_matches.extend(m_sit)
        all_unmatched_mps.extend(u_sit)

        state_winners = [w for w in winners if w["state"] == state]
        m_win, u_win = match_people(state_winners, named, role="prior_direkt_winner")
        all_matches.extend(m_win)
        all_unmatched_winners.extend(u_win)

        # Unique incumbent persons matched (either role)
        sit_matched_persons = _person_keys(m_sit, "incumbent_name", "incumbent_party")
        win_matched_persons = _person_keys(m_win, "incumbent_name", "incumbent_party")
        # Unique 2026 candidate persons that hit
        cand_persons = {
            (normalize_name(r["cand_name"]), r["cand_party"]) for r in m_sit + m_win
        }

        print(
            f"  sitting → 2026: {len(sit_matched_persons)}/{len(mps)} persons "
            f"({len(m_sit)} candidacy rows)"
        )
        if state_winners:
            print(
                f"  prior Direkt → 2026: {len(win_matched_persons)}/{len(state_winners)} "
                f"({len(m_win)} candidacy rows)"
            )
        print(f"  unique 2026 cand persons flagged: {len(cand_persons)}")
        print(f"  unmatched sitting: {len(u_sit)}")
        if u_win:
            print(f"  unmatched prior Direkt: {len(u_win)}")

        summary_rows.append(
            {
                "state": state,
                "sitting_mps": len(mps),
                "sitting_matched_persons": len(sit_matched_persons),
                "prior_direkt_winners": len(state_winners),
                "prior_direkt_matched_persons": len(win_matched_persons),
                "unique_2026_persons_flagged": len(cand_persons),
                "unmatched_sitting": len(u_sit),
                "unmatched_prior_direkt": len(u_win),
                "cand_2026_named": len(named),
            }
        )

    _write_csv(OUT_DIR / "sitting_mps.csv", all_mps)
    _write_csv(OUT_DIR / "matches.csv", all_matches)
    _write_csv(OUT_DIR / "unmatched_sitting.csv", all_unmatched_mps)
    _write_csv(OUT_DIR / "unmatched_prior_direkt.csv", all_unmatched_winners)
    _write_csv(OUT_DIR / "summary.csv", summary_rows)

    # Compact person-level flag for 2026 candidates (union of roles)
    flags: dict[tuple[str, str, str], dict] = {}
    for r in all_matches:
        key = (r["state"], r["cand_party"], normalize_name(r["cand_name"]))
        cur = flags.get(key)
        roles = set()
        if cur:
            roles.update(cur["roles"].split("|"))
        roles.add(r["role"])
        flags[key] = {
            "state": r["state"],
            "party": r["cand_party"],
            "name": r["cand_name"],
            "norm_name": key[2],
            "roles": "|".join(sorted(roles)),
            "is_sitting_mp": "1" if "sitting_mp" in roles else "0",
            "is_prior_direkt_winner": "1" if "prior_direkt_winner" in roles else "0",
            "match_how": r["match_how"],
        }
    _write_csv(
        OUT_DIR / "incumbent_candidates_2026.csv",
        list(flags.values()),
        fieldnames=[
            "state",
            "party",
            "name",
            "norm_name",
            "roles",
            "is_sitting_mp",
            "is_prior_direkt_winner",
            "match_how",
        ],
    )

    print(f"\nWrote outputs under {OUT_DIR}/")
    print("  sitting_mps.csv  direkt_winners.csv  matches.csv")
    print("  incumbent_candidates_2026.csv  summary.csv  unmatched_*.csv")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument(
        "--states",
        nargs="+",
        default=["BE", "MV", "ST"],
        choices=["BE", "MV", "ST"],
    )
    ap.add_argument(
        "--offline",
        action="store_true",
        help="Reuse wiki HTML cached under tmp/incumbents/wiki_cache/",
    )
    args = ap.parse_args()
    run(args.states, offline=args.offline)


if __name__ == "__main__":
    main()
