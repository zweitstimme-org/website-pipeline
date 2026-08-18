#!/usr/bin/env python3
"""Landtagswahl election-night nowcast for ST and MV (MVP).

Replay: LTW 2021 with L1 = 2016 Wahlbezirke (matched by AGS + WB-Nr).
Night update: shrunk global surprise on open WBs (Berlin-style, simplified).
Writes:
  output/wahlabend_nowcast_st.json
  output/wahlabend_nowcast_mv.json
"""

from __future__ import annotations

import argparse
import json
import math
import re
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
import openpyxl

REPO = Path(__file__).resolve().parents[1]
OUT_DIR = REPO / "output"

PARTIES = ("spd", "cdu", "gruene", "linke", "afd", "fdp", "others")
MAIN = ("spd", "cdu", "gruene", "linke", "afd", "fdp")
EPS = 1e-12
RNG_SEED = 42

PARTY_LABELS = {
    "spd": "SPD",
    "cdu": "CDU",
    "gruene": "GRÜNE",
    "linke": "Linke",
    "afd": "AfD",
    "fdp": "FDP",
    "others": "Sonstige",
}

# Column maps: display name fragment → code
_NAME_TO_CODE = {
    "spd": "spd",
    "cdu": "cdu",
    "afd": "afd",
    "fdp": "fdp",
    "grüne": "gruene",
    "gruene": "gruene",
    "linke": "linke",
    "die linke": "linke",
}


def _norm_party(name: str) -> str | None:
    if not name:
        return None
    s = re.sub(r"\s+", " ", str(name).replace("\n", " ")).strip().lower()
    s = s.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue")
    for key, code in _NAME_TO_CODE.items():
        if key in s:
            return code
    return None


def _wbz_key(wbz) -> str:
    """Normalize Wahlbezirk number so 1 and 000001 match."""
    if wbz is None:
        return ""
    s = str(wbz).strip()
    try:
        return str(int(float(s)))
    except ValueError:
        return s.lstrip("0") or "0"


def _ags_key(ags) -> str:
    if ags is None:
        return ""
    if isinstance(ags, float):
        return str(int(ags))
    s = str(ags).strip()
    try:
        return str(int(float(s)))
    except ValueError:
        return s


def _num(x) -> float:
    if x is None or x == "" or x in ("-", "x", "X", "."):
        return 0.0
    if isinstance(x, (int, float)):
        return float(x) if math.isfinite(float(x)) else 0.0
    s = str(x).strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _shares(counts: dict[str, float]) -> dict[str, float]:
    tot = sum(counts.values()) + EPS
    return {p: counts[p] / tot for p in PARTIES}


def _pct(sh: dict[str, float]) -> dict[str, float]:
    return {p: round(sh[p] * 100.0, 2) for p in PARTIES}


def _agg(precincts: dict[str, dict], keys: list[str] | None = None) -> dict[str, float]:
    keys = keys if keys is not None else list(precincts)
    c = {p: 0.0 for p in PARTIES}
    for a in keys:
        g = float(precincts[a]["gueltig"])
        for p in PARTIES:
            c[p] += g * precincts[a]["shares"][p]
    return _shares(c)


def _leader(sh: dict[str, float]) -> str:
    return max(MAIN, key=lambda p: sh.get(p, 0.0))


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------


def load_st_wbz(path: Path, *, year: int) -> dict[str, dict]:
    """ST Wahlbezirk end results (multi-row header)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    # Find header party row (contains 'CDU' and 'AfD')
    header_i = None
    for i, r in enumerate(rows[:15]):
        vals = [str(x) for x in (r or []) if x is not None]
        joined = " ".join(vals)
        if "CDU" in joined and "AfD" in joined and "Gültige" in joined.replace("\n", " "):
            header_i = i
            break
    if header_i is None:
        raise RuntimeError(f"ST header not found in {path}")

    header = list(rows[header_i])
    # Zweit block starts at second 'Gültige' / CDU after Erst
    erst_cdu = next(i for i, v in enumerate(header) if v and "CDU" in str(v))
    zweit_cdu = next(
        i for i, v in enumerate(header) if i > erst_cdu + 5 and v and str(v).strip().startswith("CDU")
    )
    zweit_gueltig = zweit_cdu - 1

    party_cols: dict[str, int] = {}
    for i in range(zweit_cdu, len(header)):
        code = _norm_party(header[i])
        if code and code not in party_cols and code != "others":
            party_cols[code] = i

    # Data starts after 2-3 subheader rows
    start = header_i + 1
    while start < len(rows) and (
        rows[start] is None
        or rows[start][0] in (None, "Nr.", "Nummer")
        or (isinstance(rows[start][0], str) and not str(rows[start][0]).strip()[:1].isdigit())
    ):
        # still header-ish if first cell is Nr.
        if rows[start] and str(rows[start][0]).strip() in ("Nr.", "Nummer"):
            start += 1
            continue
        if rows[start] and rows[start][0] is None and all(
            x is None or (isinstance(x, str) and "Nr" in str(x)) for x in rows[start][:5]
        ):
            start += 1
            continue
        break

    out: dict[str, dict] = {}
    for r in rows[start:]:
        if not r or r[0] is None:
            continue
        wkr_raw = r[0]
        try:
            wkr = int(str(wkr_raw).strip())
        except ValueError:
            continue
        if year >= 2021:
            ags = r[4]
            wbz = r[8]
            art = str(r[10] or "U").strip().upper()[:1] or "U"
            name = str(r[9] or "").strip()
            wber = _num(r[11])
            waehler = _num(r[15])
            gueltig = _num(r[zweit_gueltig])
        else:
            # 2016: wkr, name, ags, gem, vg…, wbz, wbz name, Urne/Brief, wber, waehler
            ags = r[2]
            wbz = r[8]
            art_s = str(r[10] or "").lower()
            art = "B" if "brief" in art_s else "U"
            name = str(r[9] or "").strip()
            wber = _num(r[11])
            waehler = _num(r[12])
            gueltig = _num(r[zweit_gueltig]) if zweit_gueltig < len(r) else 0.0

        if ags is None or wbz is None:
            continue
        ags_s = _ags_key(ags)
        wbz_s = _wbz_key(wbz)
        uid = f"{ags_s}|{wbz_s}|{art}"

        counts = {p: 0.0 for p in PARTIES}
        named = 0.0
        for p, col in party_cols.items():
            if col < len(r):
                v = _num(r[col])
                counts[p] = v
                named += v
        counts["others"] = max(0.0, gueltig - named)
        if gueltig <= 0 and named > 0:
            gueltig = named
        if gueltig <= 0:
            continue
        out[uid] = {
            "id": uid,
            "wkr": str(wkr),
            "ags": ags_s,
            "art": art,
            "name": name,
            "wber": wber,
            "waehler": waehler,
            "gueltig": gueltig,
            "counts": counts,
            "shares": _shares(counts),
        }
    return out


def load_mv_wbz(path: Path, *, year: int) -> dict[str, dict]:
    """MV 'nach Wahlbezirken' sheet."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["nach Wahlbezirken"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Party name row: contains SPD and CDU and ungültig
    header_i = None
    for i, r in enumerate(rows[:20]):
        vals = [str(x) for x in (r or []) if x is not None]
        if any("SPD" == str(x).strip() for x in (r or []) if x) and any(
            str(x).strip() == "CDU" for x in (r or []) if x
        ):
            header_i = i
            break
    if header_i is None:
        raise RuntimeError(f"MV party header not found in {path}")
    header = list(rows[header_i])

    # Erst parties after first 'gültig', Zweit after second
    gueltig_idxs = [i for i, v in enumerate(header) if v and "gültig" in str(v).lower()]
    if len(gueltig_idxs) < 2:
        raise RuntimeError(f"MV expected 2 gültig cols in {path}")
    erst_g, zweit_g = gueltig_idxs[0], gueltig_idxs[1]

    party_cols: dict[str, int] = {}
    for i in range(zweit_g + 1, len(header)):
        code = _norm_party(header[i])
        if code and code not in party_cols and code != "others":
            party_cols[code] = i

    # Data: first numeric wkr in col 0 after header+blank+code row
    start = header_i + 1
    while start < len(rows):
        r = rows[start]
        if r and r[0] is not None:
            try:
                int(r[0])
                break
            except (TypeError, ValueError):
                pass
        start += 1

    out: dict[str, dict] = {}
    for r in rows[start:]:
        if not r or r[0] is None:
            continue
        try:
            wkr = int(r[0])
        except (TypeError, ValueError):
            continue
        ags = r[2]
        wbz = r[4]
        if ags is None or wbz is None:
            continue
        ags_s = _ags_key(ags)
        wbz_n = int(float(wbz)) if not isinstance(wbz, str) else int(str(wbz).strip())
        art = "B" if wbz_n > 900 else "U"
        uid = f"{ags_s}|{wbz_n}|{art}"
        wber = _num(r[8])
        waehler = _num(r[9])
        gueltig = _num(r[zweit_g])
        counts = {p: 0.0 for p in PARTIES}
        named = 0.0
        for p, col in party_cols.items():
            if col < len(r):
                v = _num(r[col])
                counts[p] = v
                named += v
        counts["others"] = max(0.0, gueltig - named)
        if gueltig <= 0 and named > 0:
            gueltig = named
        if gueltig <= 0:
            continue
        out[uid] = {
            "id": uid,
            "wkr": str(wkr),
            "ags": ags_s,
            "art": art,
            "name": str(r[3] or "").strip(),
            "wber": wber,
            "waehler": waehler,
            "gueltig": gueltig,
            "counts": counts,
            "shares": _shares(counts),
        }
    return out


def match_l1(
    truth: dict[str, dict], l1: dict[str, dict]
) -> tuple[dict[str, dict], dict[str, dict], dict]:
    """Keep WBs present in both; build meta."""
    common = sorted(set(truth) & set(l1))
    t = {k: truth[k] for k in common}
    b = {k: l1[k] for k in common}
    meta = {
        "n_truth": len(truth),
        "n_l1": len(l1),
        "n_matched": len(common),
        "match_rate": round(len(common) / max(len(truth), 1), 3),
    }
    return t, b, meta


def build_priors(
    precincts: dict[str, dict], l1: dict[str, dict], target: dict[str, float]
) -> dict[str, dict[str, float]]:
    """Proportional swing from L1 precinct shares toward statewide target."""
    l1_land = _agg(l1)
    out: dict[str, dict[str, float]] = {}
    for a, row in precincts.items():
        base = l1[a]["shares"]
        swung = {}
        for p in PARTIES:
            if l1_land[p] > EPS:
                swung[p] = base[p] * (target[p] / l1_land[p])
            else:
                swung[p] = target[p]
        s = sum(swung.values()) + EPS
        out[a] = {p: swung[p] / s for p in PARTIES}
    return out


def nowcast_step(
    reported: set[str],
    precincts: dict[str, dict],
    prior: dict[str, dict[str, float]],
    *,
    half_life: float = 0.12,
) -> tuple[dict[str, dict[str, float]], dict]:
    all_addrs = list(precincts)
    total_g = sum(precincts[a]["gueltig"] for a in all_addrs) + EPS
    reported_g = sum(precincts[a]["gueltig"] for a in reported)
    frac = reported_g / total_g
    w = frac / (frac + half_life)

    if not reported:
        surprise = {p: 0.0 for p in PARTIES}
        nc = {a: dict(prior[a]) for a in all_addrs}
    else:
        obs = _agg(precincts, list(reported))
        tw = sum(precincts[a]["gueltig"] for a in reported) + EPS
        pri_r = {p: 0.0 for p in PARTIES}
        for a in reported:
            g = precincts[a]["gueltig"]
            for p in PARTIES:
                pri_r[p] += g * prior[a][p]
        pri_r = {p: pri_r[p] / tw for p in PARTIES}
        surprise = {p: obs[p] - pri_r[p] for p in PARTIES}
        nc = {}
        for a in all_addrs:
            if a in reported:
                nc[a] = dict(precincts[a]["shares"])
            else:
                raw = {p: prior[a][p] + w * surprise[p] for p in PARTIES}
                raw = {p: max(0.0, v) for p, v in raw.items()}
                s = sum(raw.values()) + EPS
                nc[a] = {p: raw[p] / s for p in PARTIES}

    open_frac = 1.0 - frac
    unc = {
        p: round(max(0.3, 4.0 * open_frac + abs(surprise.get(p, 0.0)) * 100 * 0.25), 2)
        for p in PARTIES
    }
    if open_frac <= 1e-9:
        unc = {p: 0.0 for p in PARTIES}
    diag = {
        "learn_weight": round(w, 4),
        "frac_votes": round(frac, 4),
        "surprise": {p: round(surprise[p] * 100, 3) for p in PARTIES},
        "uncertainty": unc,
    }
    return nc, diag


def nowcast_turnout(
    reported: set[str], precincts: dict[str, dict], l1: dict[str, dict]
) -> dict:
    addrs = list(precincts)
    total_wber = sum(float(precincts[a].get("wber") or 0.0) for a in addrs) + EPS

    def exp_voters(a: str) -> float:
        wber = float(precincts[a].get("wber") or 0.0)
        l1w = float(l1[a].get("waehler") or 0.0)
        l1b = float(l1[a].get("wber") or 0.0)
        if wber > 0 and l1b > 0 and l1w > 0:
            return wber * min(1.0, l1w / l1b)
        return max(l1w, 0.0)

    exp = {a: exp_voters(a) for a in addrs}
    truth_v = {a: float(precincts[a].get("waehler") or 0.0) for a in addrs}
    prior_city = sum(exp.values()) / total_wber
    truth_city = sum(truth_v.values()) / total_wber
    rep = [a for a in addrs if a in reported]
    open_ = [a for a in addrs if a not in reported]
    total_exp = sum(exp.values()) + EPS
    frac = sum(exp[a] for a in rep) / total_exp
    if rep:
        obs = sum(truth_v[a] for a in rep)
        er = sum(exp[a] for a in rep) + EPS
        factor_raw = obs / er
        w = frac / (frac + 0.08)
        factor = 1.0 + (factor_raw - 1.0) * w
        hat = obs + factor * sum(exp[a] for a in open_)
        nc = hat / total_wber
        naive = (obs / er) * prior_city
    else:
        nc = prior_city
        naive = prior_city
    open_w = max(0.0, 1.0 - frac)
    unc = 0.0 if open_w <= 1e-9 else round(5.0 * open_w, 2)
    return {
        "nowcast": round(float(np.clip(nc, 0, 1)) * 100, 2),
        "naive": round(float(np.clip(naive, 0, 1.5)) * 100, 2),
        "prior": round(float(np.clip(prior_city, 0, 1)) * 100, 2),
        "truth": round(float(np.clip(truth_city, 0, 1)) * 100, 2),
        "uncertainty": unc,
        "frac_wber_reported": round(frac, 4),
        "abs_err": round(abs(nc - truth_city) * 100, 2),
    }


def mae(a: dict[str, float], b: dict[str, float]) -> float:
    return float(np.mean([abs(a[p] - b[p]) for p in MAIN]))


def listen_roster_ltw(max_pos: int = 45) -> dict[str, dict]:
    """Placeholder Landeslisten for ST/MV replay (no person names)."""
    out: dict[str, dict] = {}
    for p in MAIN:
        out[p] = {
            "list_type": "landes",
            "landes": [
                {
                    "pos": i,
                    "name": f"{PARTY_LABELS[p]} · Listenplatz {i}",
                    "ph": True,
                }
                for i in range(1, max_pos + 1)
            ],
            "bezirk": {},
        }
    return out


def night_entry_mc_ltw(
    nc_land_pct: dict[str, float],
    unc_pp: dict[str, float],
    wkr_leaders: dict[str, str],
    rng: np.random.Generator,
    state_code: str,
    n_draws: int = 48,
) -> dict:
    """MC seat / list-entry bands for LTW (Landeslisten only, no Bezirkslisten)."""
    from parliament_size_sim import allocate_mv, allocate_st

    allocate = allocate_mv if state_code == "mv" else allocate_st
    directs = {p: 0 for p in MAIN}
    for lead in wkr_leaders.values():
        if lead in directs:
            directs[lead] += 1

    sizes: list[int] = []
    seats_acc: dict[str, list[int]] = {p: [] for p in MAIN}
    list_land_acc: dict[str, list[int]] = {p: [] for p in MAIN}
    for _ in range(n_draws):
        draw = {p: 0.0 for p in PARTIES}
        for p in PARTIES:
            sd = float(unc_pp.get(p, 0.0)) / 1.28
            draw[p] = max(0.0, float(nc_land_pct.get(p, 0.0)) + rng.normal(0.0, sd))
        tot = sum(draw.values()) or 1.0
        frac = {p: draw[p] / tot for p in PARTIES}
        alloc = allocate(frac, directs)
        sizes.append(int(alloc["size"]))
        for p in MAIN:
            s_p = int(alloc["seats"].get(p, 0))
            seats_acc[p].append(s_p)
            list_land_acc[p].append(max(0, s_p - directs.get(p, 0)))

    def q(vals: list[int]) -> list[int]:
        arr = np.asarray(vals)
        return [
            int(np.percentile(arr, 10)),
            int(np.percentile(arr, 50)),
            int(np.percentile(arr, 90)),
        ]

    return {
        "n_draws": n_draws,
        "size": q(sizes),
        "seats": {p: q(seats_acc[p]) for p in MAIN},
        "directs": directs,
        "list_seats": {p: q(list_land_acc[p]) for p in MAIN},
    }


def run_state(state: str, *, n_steps: int = 40) -> dict:
    state = state.lower()
    if state == "st":
        raw = REPO / "sachsen-anhalt" / "wahlabend" / "raw"
        truth = load_st_wbz(raw / "LT2021_WBZ.xlsx", year=2021)
        l1_all = load_st_wbz(raw / "LT2016_WBZ.xlsx", year=2016)
        label = "Sachsen-Anhalt"
        election = "LTW2021"
        n_wkr = 41
    elif state == "mv":
        raw = REPO / "mecklenburg-vorpommern" / "wahlabend" / "raw"
        truth = load_mv_wbz(raw / "LTW2021_WBZ.xlsx", year=2021)
        l1_all = load_mv_wbz(raw / "LTW2016_WBZ.xlsx", year=2016)
        label = "Mecklenburg-Vorpommern"
        election = "LTW2021"
        n_wkr = 36
    else:
        raise ValueError(state)

    precincts, l1, match_meta = match_l1(truth, l1_all)
    addrs = sorted(precincts)
    truth_land = _agg(precincts)
    # Replay prior = L1 land (honest: night must learn the swing)
    prior_target = _agg(l1)
    prior = build_priors(precincts, l1, prior_target)
    prior_land = _agg({a: {"gueltig": precincts[a]["gueltig"], "shares": prior[a]} for a in addrs})

    # Geography
    by_wkr: dict[str, list[str]] = {}
    for a, r in precincts.items():
        by_wkr.setdefault(r["wkr"], []).append(a)

    rng = np.random.default_rng(RNG_SEED)
    order = list(addrs)
    rng.shuffle(order)

    total_g = sum(precincts[a]["gueltig"] for a in addrs)
    steps = []
    reported: set[str] = set()
    reported_g = 0.0
    targets = np.linspace(0.0, 1.0, n_steps)
    cursor = 0
    prev_unc = None
    mc_rng = np.random.default_rng(RNG_SEED + (11 if state == "st" else 12))

    for tfrac in targets:
        while cursor < len(order) and (reported_g / (total_g + EPS)) < tfrac - 1e-12:
            a = order[cursor]
            reported.add(a)
            reported_g += precincts[a]["gueltig"]
            cursor += 1
        if tfrac >= 1.0 - 1e-12:
            reported = set(addrs)
            reported_g = total_g

        nc, diag = nowcast_step(reported, precincts, prior)
        unc = diag["uncertainty"]
        if prev_unc:
            unc = {p: round(min(prev_unc[p], unc[p]), 2) for p in PARTIES}
        prev_unc = unc
        diag["uncertainty"] = unc

        nc_land = _agg({a: {"gueltig": precincts[a]["gueltig"], "shares": nc[a]} for a in addrs})
        if reported:
            naive_land = _agg(precincts, list(reported))
        else:
            naive_land = dict(prior_land)

        # WK regions
        wkr_out = {}
        for wid, ua in sorted(by_wkr.items(), key=lambda kv: int(kv[0])):
            now = _agg({a: {"gueltig": precincts[a]["gueltig"], "shares": nc[a]} for a in ua}, ua)
            tru = _agg(precincts, ua)
            rep_here = [a for a in ua if a in reported]
            frac_w = sum(precincts[a]["gueltig"] for a in rep_here) / (
                sum(precincts[a]["gueltig"] for a in ua) + EPS
            )
            lp = _leader(now)
            lt = _leader(tru)
            top2 = sorted(MAIN, key=lambda p: now[p], reverse=True)[:2]
            margin = (now[top2[0]] - now[top2[1]]) * 100
            # crude p_lead from margin vs unc
            m_u = math.sqrt(unc.get(top2[0], 1) ** 2 + unc.get(top2[1], 1) ** 2)
            z = margin / (m_u + 0.5)
            p_lead = float(0.5 * (1 + math.erf(z / math.sqrt(2))))
            complete = frac_w >= 0.999
            likely = p_lead >= 0.90
            called = p_lead >= 0.999 and frac_w > 0 and (
                complete or margin >= (1 - frac_w) * 40
            )
            wkr_out[wid] = {
                "frac_reported": round(frac_w, 4),
                "n_reported": len(rep_here),
                "nowcast": _pct(now),
                "truth": _pct(tru),
                "leader_pred": lp,
                "leader_truth": lt,
                "leader_ok": lp == lt,
                "direct_pred": lp,
                "margin": round(margin, 2),
                "p_lead": round(p_lead, 4),
                "likely": likely,
                "called": called,
                "complete": complete,
                "uncertainty": {p: unc[p] for p in PARTIES},
                "ballot": "erst_proxy",
            }

        turnout = nowcast_turnout(reported, precincts, l1)
        entry_mc = night_entry_mc_ltw(
            _pct(nc_land),
            unc,
            {wid: r["direct_pred"] for wid, r in wkr_out.items()},
            mc_rng,
            state,
        )
        frac_now = round(reported_g / (total_g + EPS), 4)
        steps.append(
            {
                "frac_reported": frac_now,
                "n_reported": len(reported),
                "n_total": len(addrs),
                "clock": None,
                "clock_source": "sim",
                "nowcast": _pct(nc_land),
                "naive": _pct(naive_land),
                "prior": _pct(prior_land),
                "truth": _pct(truth_land),
                "mae_nowcast": round(mae(nc_land, truth_land) * 100, 4),
                "mae_naive": round(mae(naive_land, truth_land) * 100, 4),
                "learn_weight": diag["learn_weight"],
                "representativeness": None,
                "surprise": diag["surprise"],
                "uncertainty": unc,
                "turnout": turnout,
                "by_wkr": wkr_out,
                "by_bezirk": {},
                "entry_mc": entry_mc,
                "scenario_probs": None,
                "eval": None,
            }
        )

    # pick mid-night MAE for summary
    mid = steps[len(steps) // 4]
    payload = {
        "election": election,
        "state": state,
        "state_label": label,
        "baseline": "L1 = LTW 2016 matched WB; π₀ = L1-Land (Replay)",
        "parties": list(PARTIES),
        "party_labels": PARTY_LABELS,
        "n_precincts": len(addrs),
        "n_wkr": n_wkr,
        "match": match_meta,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "model": {
            "description": (
                f"{label} Wahlabend-Nowcast (MVP): global Surprise × Lerngewicht; "
                "Direkt = Zweit-Proxy je WK bis Erst-Loader folgt."
            )
        },
        "call_threshold": 0.90,
        "hard_call_threshold": 0.999,
        "geo_units": {
            "land": [{"id": state.upper(), "label": label}],
            "bezirk": [],
            "wkr": [
                {"id": wid, "label": f"WK {wid}", "bezirk": None}
                for wid in sorted(by_wkr, key=int)
            ],
        },
        "scenarios": {
            "random": {
                "label": "Zufällige Meldeordnung",
                "steps": steps,
                "wkr_calls": {},
                "summary": {
                    "mae_at_25": mid["mae_nowcast"],
                    "mae_naive_at_25": mid["mae_naive"],
                },
            }
        },
        "scenario": "random",
        "features": {
            "bezirkslisten": False,
            "listen_einzug": True,
        },
        "listen_mode": "landes",
        "listen_roster_2026": listen_roster_ltw(),
        "listen_roster_note": (
            "Platzhalter-Listenplätze (Replay LTW 2021). Nur Landeslisten — "
            "keine Bezirkslisten wie in Berlin. Einzug = Nowcast-Sitze; "
            "Wackelbereich = MC p10–p90."
        ),
        "precincts": [
            {
                "id": a,
                "wkr": precincts[a]["wkr"],
                "art": precincts[a]["art"],
                "name": precincts[a]["name"],
                "bezirk": None,
                "gueltig": int(round(precincts[a]["gueltig"])),
                "wber": int(round(precincts[a].get("wber") or 0)),
                "waehler": int(round(precincts[a].get("waehler") or 0)),
                "counts": {
                    p: int(round(precincts[a]["counts"][p])) for p in PARTIES
                },
            }
            for a in addrs
        ],
    }
    return payload


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--states", default="st,mv")
    ap.add_argument("--steps", type=int, default=40)
    args = ap.parse_args()
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    for st in [s.strip() for s in args.states.split(",") if s.strip()]:
        print(f"Running {st} …")
        payload = run_state(st, n_steps=args.steps)
        out = OUT_DIR / f"wahlabend_nowcast_{st}.json"
        out.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
        sc = payload["scenarios"]["random"]["summary"]
        print(
            f"  wrote {out}  matched={payload['match']['n_matched']}/"
            f"{payload['match']['n_truth']}  MAE@~25%={sc['mae_at_25']} "
            f"(naive {sc['mae_naive_at_25']})"
        )


if __name__ == "__main__":
    main()
