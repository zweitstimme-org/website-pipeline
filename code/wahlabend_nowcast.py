#!/usr/bin/env python3
"""Berlin AGH election-night nowcast (MVP).

Prior: proportional swing from L1 (AGH 2016→2023 precincts) toward a statewide
pre-election target π₀. Night update: compare reported results to what the prior
predicted *for those same precincts*, treat the gap as a swing surprise, and
extrapolate only as far as the reported set is informative (vote share ×
composition representativeness vs the full city).

Evaluation: simulated reporting on AGH 2023. Writes
  output/wahlabend_nowcast_replay.json for the preview timeline UI.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

TZ_BERLIN = ZoneInfo("Europe/Berlin")

import numpy as np
import openpyxl

REPO = Path(__file__).resolve().parents[1]
RAW = REPO / "berlin" / "wahlabend" / "raw"
PROCESSED = REPO / "berlin" / "wahlabend" / "processed"
OUT = REPO / "output" / "wahlabend_nowcast_replay.json"
PRIOR_2023 = PROCESSED / "prior_be_agh2023.json"
PRIOR_2016 = PROCESSED / "prior_be_agh2016.json"

PARTIES = ("spd", "cdu", "gruene", "linke", "afd", "fdp", "others")
MAIN_PARTIES = ("spd", "cdu", "gruene", "linke", "afd", "fdp")
HURDLE = 0.05
AGH_BASE_SEATS = 130
# Berlin 2026: CDU/SPD/Linke treten mit Bezirkslisten an, Rest Landesliste.
BE_BEZIRKSLISTE_PARTIES = ("spd", "cdu", "linke")
# „Wahrscheinlich“ ab P(Führung hält) ≥ 90 % — auch ohne lokale Auszählung
# (Prior/Landestrend reicht; harter Call braucht dagegen lokale Meldung).
CALL_THRESHOLD = 0.90
# Harter Call ab 99,9 % + lokale Meldung + Restanteil kann die Marge
# nicht mehr kippen (sonst wäre 100 % Auszählung kein Call, nur Abwarten).
HARD_CALL_THRESHOLD = 0.999
# Konservativer Rest-Swing (PP) × offener Stimmenanteil → max. Flip-Marge.
CALL_RESIDUAL_SWING_PP = 40.0
# Ohne lokale Meldung: Prior-Race-Floor (pp) gegen überconfidentes P(Führung).
# AGH2023: Prior-Anteil RMSE ≈ 4 PP; sonst schrumpft das Band mit dem Landestrend,
# während der WK-Punkt nahe am Prior bleibt (z. B. WK 62 Linke→CDU).
WK_OPEN_PRIOR_FLOOR_PP = 3.5
PARTY_LABELS = {
    "spd": "SPD",
    "cdu": "CDU",
    "gruene": "GRÜNE",
    "linke": "Linke",
    "afd": "AfD",
    "fdp": "FDP",
    "others": "Sonstige",
}
# Excel column names → code
XLSX_PARTY_MAP = {
    "SPD": "spd",
    "CDU": "cdu",
    "GRÜNE": "gruene",
    "DIE LINKE": "linke",
    "AfD": "afd",
    "FDP": "fdp",
}

# Structure percent fields used for similarity (2023 Strukturdaten)
STRUCT_FEATURES = [
    "Einwohner 65 und älter Prozent",
    "Ausländer Prozent",
    "Deutsche 18 - 25 Prozent",
    "Deutsche 25 - 35 Prozent",
    "Deutsche 70+ Prozent",
    "Deutsche 18+ Migrationshintergrund Prozent",
    "Deutsche 18+ Familienstand ledig Prozent",
]

EPS = 1e-9
RNG_SEED = 20260807


def _f(x) -> float:
    if x is None or x == "":
        return 0.0
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _shares_from_counts(counts: dict[str, float]) -> dict[str, float]:
    tot = sum(counts.values())
    if tot <= 0:
        return {p: 1.0 / len(PARTIES) for p in PARTIES}
    return {p: counts[p] / tot for p in PARTIES}


def _renorm(shares: dict[str, float]) -> dict[str, float]:
    clipped = {p: max(0.0, float(shares.get(p, 0.0))) for p in PARTIES}
    s = sum(clipped.values())
    if s <= 0:
        return {p: 1.0 / len(PARTIES) for p in PARTIES}
    return {p: clipped[p] / s for p in PARTIES}


def _read_xlsx_sheet(path: Path, sheet: str | None = None):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet is None:
        sheet = next(
            s
            for s in wb.sheetnames
            if not s.lower().startswith("impress") and s.lower() not in ("erläuterungen", "titel")
        )
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header = [str(c).replace("\n", "") if c is not None else "" for c in rows[0]]
    out = []
    for row in rows[1:]:
        d = {header[i]: row[i] for i in range(len(header)) if header[i]}
        if d.get("Adresse"):
            out.append(d)
    return out


def load_truth_2023() -> dict[str, dict]:
    """AGH 2023 Zweitstimme by Wahlbezirk."""
    return _load_agh_precinct_sheet("AGH_W2")


def load_truth_erst_2023() -> dict[str, dict]:
    """AGH 2023 Erststimme by Wahlbezirk (same Adressen as Zweit)."""
    return _load_agh_precinct_sheet("AGH_W1")


def _load_agh_precinct_sheet(sheet: str) -> dict[str, dict]:
    rows = _read_xlsx_sheet(RAW / "DL_BE_AGHBVV2023.xlsx", sheet)
    out: dict[str, dict] = {}
    for r in rows:
        addr = str(r["Adresse"]).strip()
        counts = {p: 0.0 for p in PARTIES}
        for col, code in XLSX_PARTY_MAP.items():
            counts[code] += _f(r.get(col))
        named = sum(counts[p] for p in PARTIES if p != "others")
        gueltig = _f(r.get("Gültige Stimmen"))
        counts["others"] = max(0.0, gueltig - named)
        bez = str(r.get("Bezirksnummer") or "").zfill(2)
        wkr_local = int(
            _f(
                r.get("Abgeordnetenhauswahlkreis")
                or r.get("Abgeordneten-hauswahlkreis")
            )
        )
        out[addr] = {
            "adresse": addr,
            "bezirk": bez,
            "bezirk_name": str(r.get("Bezirksname") or ""),
            "art": str(r.get("Wahlbezirksart") or "").strip().upper()[:1] or "W",
            "wkr_local": wkr_local,
            "wkr": 0,
            "ostwest": str(r.get("OstWest") or ""),
            "wber": _f(r.get("Wahlberechtigte insgesamt")),
            "waehler": _f(r.get("Wählende")),
            "gueltig": gueltig,
            "counts": counts,
            "shares": _shares_from_counts(counts),
        }
    pairs = sorted({(r["bezirk"], r["wkr_local"]) for r in out.values()})
    city = {pair: i + 1 for i, pair in enumerate(pairs)}
    for r in out.values():
        r["wkr"] = city[(r["bezirk"], r["wkr_local"])]
    return out


def load_baseline_2016_on_2023() -> dict[str, dict]:
    rows = _read_xlsx_sheet(RAW / "DL_BE_AGH2023_AH2016.xlsx", "2016_Zweitstimme")
    out: dict[str, dict] = {}
    for r in rows:
        addr = str(r["Adresse"]).strip()
        counts = {p: 0.0 for p in PARTIES}
        for col, code in XLSX_PARTY_MAP.items():
            counts[code] += _f(r.get(col))
        named = sum(counts[p] for p in PARTIES if p != "others")
        gueltig = _f(r.get("Gültige Stimmen"))
        counts["others"] = max(0.0, gueltig - named)
        out[addr] = {
            "gueltig": gueltig,
            "waehler": _f(r.get("Wählende")),
            "wber": _f(r.get("Wahlberechtigte insgesamt")),
            "counts": counts,
            "shares": _shares_from_counts(counts),
        }
    return out


def load_baseline_erst_2016(
    zweit_l1: dict[str, dict],
) -> tuple[dict[str, dict], dict]:
    """2016 Erststimme where Adresse overlaps 2023; else fall back to Zweit-L1.

    AfS has no 2016→2023 Erst-Remap; ~2400 Urnen match by Adresse, Briefwahl
    and remapped cells use Zweit-L1 as e_l1 stand-in for the OLS prior.
    """
    rows = _read_xlsx_sheet(RAW / "DL_BE_EE_WB_AH2016.xlsx", "Erststimme")
    out: dict[str, dict] = {}
    for r in rows:
        addr = str(r.get("Adresse") or "").strip()
        if not addr or addr not in zweit_l1:
            continue
        counts = {p: 0.0 for p in PARTIES}
        for col, code in XLSX_PARTY_MAP.items():
            counts[code] += _f(r.get(col))
        named = sum(counts[p] for p in PARTIES if p != "others")
        gueltig = _f(r.get("Gültige Stimmen"))
        counts["others"] = max(0.0, gueltig - named)
        out[addr] = {
            "gueltig": gueltig,
            "shares": _shares_from_counts(counts),
            "from_2016_erst": True,
        }
    n_fallback = 0
    for addr, z in zweit_l1.items():
        if addr in out:
            continue
        out[addr] = {
            "gueltig": z.get("gueltig", 0.0),
            "shares": dict(z["shares"]),
            "from_2016_erst": False,
        }
        n_fallback += 1
    meta = {
        "n_from_2016_erst": sum(1 for a, r in out.items() if r.get("from_2016_erst")),
        "n_fallback_zweit_l1": n_fallback,
    }
    return out, meta


def load_erst_ols_coef() -> tuple[np.ndarray, float]:
    """β̂ from district_model_coefs.json (resp_E ~ Z + e_l1 + no_cand)."""
    path = REPO / "data" / "district_model_coefs.json"
    if not path.exists():
        # Fallback ≈ identity on Zweit if model missing
        return np.array([0.0, 1.0, 0.0, 0.0], dtype=float), 0.03
    payload = json.loads(path.read_text(encoding="utf-8"))
    beta = np.array(payload.get("coef") or [0.0, 1.0, 0.0, 0.0], dtype=float)
    sigma = float(payload.get("sigma") or 0.03)
    return beta, sigma


def erst_shares_from_zweit(
    z_shares: dict[str, float],
    e_l1: dict[str, float],
    beta: np.ndarray,
) -> dict[str, float]:
    """Deterministic precinct Erst prior via the calibrated district OLS."""
    modeled = ("spd", "cdu", "gruene", "linke", "afd", "fdp")
    e = {p: 0.0 for p in PARTIES}
    modeled_sum = 0.0
    for p in modeled:
        no_cand = 1.0 if e_l1.get(p, 0.0) <= EPS else 0.0
        mu = (
            float(beta[0])
            + float(beta[1]) * z_shares.get(p, 0.0)
            + float(beta[2]) * e_l1.get(p, 0.0)
            + float(beta[3]) * no_cand
        )
        e[p] = max(0.0, mu)
        modeled_sum += e[p]
    if modeled_sum > 1.0:
        for p in modeled:
            e[p] /= modeled_sum
        modeled_sum = 1.0
    rem = max(0.0, 1.0 - modeled_sum)
    z_res = z_shares.get("others", 0.0)
    e["others"] = rem  # no BSW in AGH party set here
    if z_res > EPS and rem > 0:
        # keep residual mass in others (AGH 2023 party bag)
        e["others"] = rem
    return _renorm(e)


def build_erst_priors(
    zweit_priors: dict[str, dict[str, float]],
    erst_l1: dict[str, dict],
    beta: np.ndarray,
) -> dict[str, dict[str, float]]:
    out = {}
    for a, z in zweit_priors.items():
        el1 = (erst_l1.get(a) or {}).get("shares") or z
        out[a] = erst_shares_from_zweit(z, el1, beta)
    return out


def load_struktur_2023() -> dict[str, np.ndarray]:
    rows = _read_xlsx_sheet(RAW / "DL_BE_AGH2023_Strukturdaten.xlsx", "Strukturdaten")
    # Collect available feature columns
    feats = [f for f in STRUCT_FEATURES if any(f in r for r in rows[:1]) or True]
    # verify against header keys of first row
    keys = set(rows[0]) if rows else set()
    feats = [f for f in STRUCT_FEATURES if f in keys]
    if not feats:
        # fallback: any Prozent column
        feats = sorted(k for k in keys if "Prozent" in k)[:8]

    mat: dict[str, np.ndarray] = {}
    raw_vecs = []
    addrs = []
    for r in rows:
        addr = str(r["Adresse"]).strip()
        vec = np.array([_f(r.get(f)) for f in feats], dtype=float)
        addrs.append(addr)
        raw_vecs.append(vec)
    if not raw_vecs:
        return {}
    X = np.vstack(raw_vecs)
    # z-score columns
    mu = X.mean(axis=0)
    sd = X.std(axis=0)
    sd = np.where(sd < 1e-6, 1.0, sd)
    Z = (X - mu) / sd
    for i, addr in enumerate(addrs):
        mat[addr] = Z[i]
    return mat


@dataclass
class NowcastConfig:
    k_neighbors: int = 40
    mix_local: float = 0.55
    shrink_half_life_votes: float = 0.10
    representativeness_power: float = 1.25


# forecast_state_*.json party_code → nowcast code
_FORECAST_PARTY_ALIAS = {
    "spd": "spd",
    "cdu": "cdu",
    "gru": "gruene",
    "gruene": "gruene",
    "lin": "linke",
    "linke": "linke",
    "afd": "afd",
    "fdp": "fdp",
    "oth": "others",
    "others": "others",
    "bsw": "others",
}


def load_prior_target(path: Path) -> tuple[dict[str, float], dict]:
    """Load statewide π₀ shares (fractions) from prior JSON."""
    payload = json.loads(path.read_text(encoding="utf-8"))
    shares = payload.get("shares") or {}
    out = {p: float(shares.get(p, 0.0)) for p in PARTIES}
    s = sum(out.values())
    if s <= 0:
        raise ValueError(f"Empty prior shares in {path}")
    if s > 1.5:
        out = {p: out[p] / 100.0 for p in PARTIES}
    return _renorm(out), payload


def prior_uncertainty_pp(prior_meta: dict | None) -> dict[str, float]:
    """Party-specific soft-80% half-widths (pp) from the statewide forecast/prior.

    Prefers ``uncertainty_pp`` on the prior payload; else derives from a
    ``forecast_state`` parties[] list (high−low)/2. Fallback 2.3 pp flat.
    """
    fallback = 2.3
    meta = prior_meta or {}
    raw = meta.get("uncertainty_pp") or {}
    out = {p: float(raw[p]) for p in PARTIES if p in raw and raw[p] is not None}
    if len(out) == len(PARTIES):
        return {p: round(max(0.5, out[p]), 2) for p in PARTIES}

    # Optional: parties from forecast_state_*.json embedded or side-loaded
    parties = meta.get("parties") or []
    from_fcst: dict[str, float] = {}
    for row in parties:
        if not isinstance(row, dict):
            continue
        code = _FORECAST_PARTY_ALIAS.get(
            str(row.get("party_code") or row.get("party") or "").lower()
        )
        if not code or code not in PARTIES:
            continue
        try:
            lo = float(row["low"])
            hi = float(row["high"])
        except (KeyError, TypeError, ValueError):
            continue
        # Merge BSW into others: keep max half-width if both present
        half = abs(hi - lo) / 2.0
        from_fcst[code] = max(from_fcst.get(code, 0.0), half)
    if from_fcst:
        return {
            p: round(max(0.5, float(from_fcst.get(p, fallback))), 2) for p in PARTIES
        }
    return {p: fallback for p in PARTIES}

def proportional_swing_shares(
    district_l1: dict[str, float],
    state_l1: dict[str, float],
    state_new: dict[str, float],
) -> dict[str, float]:
    out = {}
    for p in PARTIES:
        l1 = state_l1[p]
        if l1 > EPS:
            out[p] = district_l1[p] * (1.0 + (state_new[p] - l1) / l1)
        else:
            out[p] = state_new[p]
    return _renorm(out)


def build_priors(
    precincts: dict[str, dict],
    l1: dict[str, dict],
    *,
    prior_target: dict[str, float],
) -> dict[str, dict[str, float]]:
    addrs = sorted(set(precincts) & set(l1))
    l1_land = aggregate_shares(precincts, {a: l1[a]["shares"] for a in addrs})
    return {
        a: proportional_swing_shares(l1[a]["shares"], l1_land, prior_target)
        for a in addrs
    }


def composition_representativeness(
    reported: set[str],
    precincts: dict[str, dict],
    l1_shares: dict[str, dict[str, float]],
    struktur: dict[str, np.ndarray],
) -> float:
    """1 = reported set mirrors city L1 party mix (+ structure); 0 = alien sample."""
    if not reported:
        return 0.0
    all_addrs = list(l1_shares)
    full = aggregate_shares(precincts, l1_shares, all_addrs)
    rep = aggregate_shares(precincts, l1_shares, list(reported))
    # Total variation distance on party shares → similarity
    tv = 0.5 * sum(abs(full[p] - rep[p]) for p in PARTIES)
    party_sim = max(0.0, 1.0 - tv / 0.25)  # TV≥0.25 ⇒ ~0

    # Structure mean distance (if available)
    feat_rep = [struktur[a] for a in reported if a in struktur]
    feat_all = [struktur[a] for a in all_addrs if a in struktur]
    if feat_rep and feat_all:
        mu_r = np.mean(np.vstack(feat_rep), axis=0)
        mu_a = np.mean(np.vstack(feat_all), axis=0)
        dist = float(np.linalg.norm(mu_r - mu_a))
        struct_sim = float(np.exp(-0.5 * dist))
    else:
        struct_sim = party_sim

    # Urne/Brief mix
    def art_frac(addrs):
        g_w = sum(precincts[a]["gueltig"] for a in addrs if precincts[a]["art"] == "W")
        g = sum(precincts[a]["gueltig"] for a in addrs) + EPS
        return g_w / g

    art_diff = abs(art_frac(all_addrs) - art_frac(list(reported)))
    art_sim = max(0.0, 1.0 - art_diff / 0.5)

    return float(np.clip(0.5 * party_sim + 0.3 * struct_sim + 0.2 * art_sim, 0.0, 1.0))


def learn_weight(
    frac_votes: float,
    representativeness: float,
    cfg: NowcastConfig,
) -> float:
    """How much to trust swing *corrections* from the reported sample."""
    size_w = frac_votes / (frac_votes + cfg.shrink_half_life_votes)
    # Non-representative samples: down-weight learning hard
    rep_w = representativeness ** cfg.representativeness_power
    return float(np.clip(size_w * rep_w, 0.0, 1.0))


def _waehler(row: dict) -> float:
    """Wählende; fall back to gültige Stimmen if needed."""
    w = row.get("waehler")
    if w is None or (isinstance(w, float) and np.isnan(w)):
        w = row.get("gueltig") or 0.0
    return float(w or 0.0)


def _expected_voters_l1(row: dict, l1_row: dict) -> float:
    """Prior expected voters for a WB.

    Urne: usually has Wahlberechtigte. Briefwahlbezirke often have wber=0 in
    AfS tables (eligibles sit on the paired Urne) — then use L1 Wählende as
    the capacity prior so Brief is not dropped from the citywide total.
    """
    wber = float(row.get("wber") or 0.0)
    l1_w = _waehler(l1_row)
    if wber > 0:
        l1_wber = float(l1_row.get("wber") or 0.0)
        if l1_wber > 0 and l1_w > 0:
            return wber * float(np.clip(l1_w / l1_wber, 0.0, 1.0))
        if l1_w > 0:
            return l1_w
        return wber * 0.6
    return max(l1_w, 0.0)


def nowcast_turnout(
    reported: set[str],
    precincts: dict[str, dict],
    l1: dict[str, dict],
    *,
    prior_unc_pp: float = 5.0,
) -> dict:
    """Citywide turnout nowcast from reported WBs + L1 voter residual.

    On election night Berlin often publishes Wahlbeteiligung alongside
    Zweitstimme before Erststimme is complete — same Meldefluss, scalar target.
    """
    addrs = list(precincts)
    total_wber = sum(float(precincts[a].get("wber") or 0.0) for a in addrs) + EPS
    if total_wber <= EPS:
        return {
            "nowcast": None,
            "naive": None,
            "prior": None,
            "truth": None,
            "uncertainty": None,
            "frac_wber_reported": 0.0,
        }

    exp_l1 = {
        a: _expected_voters_l1(precincts[a], l1.get(a) or {}) for a in addrs
    }
    truth_v = {a: _waehler(precincts[a]) for a in addrs}

    prior_city = sum(exp_l1.values()) / total_wber
    truth_city = sum(truth_v.values()) / total_wber

    reported_here = [a for a in addrs if a in reported]
    open_here = [a for a in addrs if a not in reported]
    # Progress ≈ reported expected voters (Brief has wber=0).
    total_exp = sum(exp_l1.values()) + EPS
    reported_exp = sum(exp_l1[a] for a in reported_here)
    frac_w = reported_exp / total_exp

    if reported_here:
        obs = sum(truth_v[a] for a in reported_here)
        exp_rep = sum(exp_l1[a] for a in reported_here) + EPS
        # Naive: scale city prior by reported over/under vs L1 expectation.
        naive_city = (obs / exp_rep) * prior_city
        factor_raw = obs / exp_rep
        w = frac_w / (frac_w + 0.08)
        factor = 1.0 + (factor_raw - 1.0) * w
        voters_hat = obs + factor * sum(exp_l1[a] for a in open_here)
        nowcast_city = voters_hat / total_wber
    else:
        naive_city = prior_city
        nowcast_city = prior_city

    # Band tracks open expected-voter share: full prior width at 0 %, zero at 100 %.
    # Prior ~5 PP: L1→AGH2023 was ~4.5 PP; calendar swings between elections are larger.
    open_w = max(0.0, 1.0 - frac_w)
    if open_w <= 1e-9 or frac_w >= 0.999:
        unc = 0.0
    else:
        unc = float(prior_unc_pp) * open_w
        unc = float(np.clip(unc, 0.0, 10.0))

    return {
        "nowcast": round(float(np.clip(nowcast_city, 0.0, 1.0)) * 100.0, 2),
        "naive": round(float(np.clip(naive_city, 0.0, 1.5)) * 100.0, 2),
        "prior": round(float(np.clip(prior_city, 0.0, 1.0)) * 100.0, 2),
        "truth": round(float(np.clip(truth_city, 0.0, 1.0)) * 100.0, 2),
        "uncertainty": round(unc, 2),
        "frac_wber_reported": round(frac_w, 4),
        "abs_err": round(abs(nowcast_city - truth_city) * 100.0, 2),
    }


def aggregate_shares(
    precincts: dict[str, dict],
    shares_by_addr: dict[str, dict[str, float]],
    addrs: list[str] | None = None,
) -> dict[str, float]:
    counts = {p: 0.0 for p in PARTIES}
    keys = addrs if addrs is not None else list(shares_by_addr)
    for a in keys:
        g = float(precincts[a]["gueltig"])
        sh = shares_by_addr[a]
        for p in PARTIES:
            counts[p] += g * sh[p]
    return _shares_from_counts(counts)


def _pct_shares(shares: dict[str, float]) -> dict[str, float]:
    return {p: round(shares[p] * 100, 3) for p in PARTIES}


def _geo_catalog(precincts: dict[str, dict]) -> dict:
    """Static labels for Land / Bezirk / AGH-Wahlkreis scopes."""
    bez: dict[str, str] = {}
    wkr: dict[str, dict] = {}
    for a, r in precincts.items():
        bid = str(r["bezirk"]).zfill(2)
        if bid not in bez:
            bez[bid] = r.get("bezirk_name") or f"Bezirk {bid}"
        wid = str(int(r["wkr"]))
        if wid not in wkr:
            local = int(r.get("wkr_local") or r["wkr"])
            wkr[wid] = {
                "id": wid,
                "label": f"WK {int(wid):02d} · {bez[bid]} ({local})",
                "bezirk": bid,
                "wkr_local": local,
            }
    return {
        "land": [{"id": "BE", "label": "Berlin (Landesliste)"}],
        "bezirk": [
            {"id": k, "label": f"{k} · {bez[k]}"} for k in sorted(bez)
        ],
        "wkr": [wkr[k] for k in sorted(wkr, key=lambda x: int(x))],
    }


def _precinct_index(
    precincts: dict[str, dict],
    precincts_erst: dict[str, dict] | None = None,
) -> list[dict]:
    """Compact precinct meta + vote tallies for coverage / Rohstand UI."""
    out = []
    for a in sorted(precincts):
        r = precincts[a]
        entry: dict = {
            "id": a,
            "bezirk": str(r["bezirk"]).zfill(2),
            "wkr": str(int(r["wkr"])),
            "art": r["art"],
            "name": r.get("bezirk_name") or "",
            "gueltig": int(round(float(r["gueltig"]))),
            "wber": int(round(float(r.get("wber") or 0))),
            "waehler": int(round(_waehler(r))),
            "counts": {p: int(round(r["counts"][p])) for p in PARTIES},
        }
        if precincts_erst and a in precincts_erst:
            er = precincts_erst[a]
            g_er = float(er["gueltig"])
            entry["gueltig_erst"] = int(round(g_er))
            if "counts" in er:
                entry["counts_erst"] = {
                    p: int(round(er["counts"][p])) for p in PARTIES
                }
            else:
                entry["counts_erst"] = {
                    p: int(round(g_er * er["shares"][p])) for p in PARTIES
                }
        out.append(entry)
    return out


def _addrs_by_scope(precincts: dict[str, dict]) -> dict[str, dict[str, list[str]]]:
    by_bez: dict[str, list[str]] = {}
    by_wkr: dict[str, list[str]] = {}
    for a, r in precincts.items():
        by_bez.setdefault(str(r["bezirk"]).zfill(2), []).append(a)
        by_wkr.setdefault(str(int(r["wkr"])), []).append(a)
    return {
        "land": {"BE": list(precincts.keys())},
        "bezirk": by_bez,
        "wkr": by_wkr,
    }


def _region_truth_prior(
    unit_addrs: list[str],
    precincts: dict[str, dict],
    prior_shares: dict[str, dict[str, float]],
) -> dict:
    """Static regional truth/prior (constant across the night)."""
    truth = aggregate_shares(
        precincts, {a: precincts[a]["shares"] for a in unit_addrs}, unit_addrs
    )
    prior = aggregate_shares(precincts, prior_shares, unit_addrs)
    return {
        "n_total": len(unit_addrs),
        "prior": _pct_shares(prior),
        "truth": _pct_shares(truth),
        "mae_prior": round(mae(prior, truth) * 100, 4),
    }


def _region_step(
    unit_addrs: list[str],
    precincts: dict[str, dict],
    reported: set[str],
    nc: dict[str, dict[str, float]],
    prior_shares: dict[str, dict[str, float]],
    static: dict,
    *,
    land_diag: dict | None = None,
    prior_unc_pp: dict[str, float] | None = None,
) -> dict:
    """Time-varying nowcast / naive for one geographic unit."""
    truth_sh = {p: static["truth"][p] / 100.0 for p in PARTIES}
    prior_sh = {p: static["prior"][p] / 100.0 for p in PARTIES}
    nowcast = aggregate_shares(precincts, nc, unit_addrs)
    reported_here = [a for a in unit_addrs if a in reported]
    total_g = sum(precincts[a]["gueltig"] for a in unit_addrs) + EPS
    reported_g = sum(precincts[a]["gueltig"] for a in reported_here)
    if reported_here:
        naive = aggregate_shares(
            precincts,
            {a: precincts[a]["shares"] for a in reported_here},
            reported_here,
        )
    else:
        naive = dict(prior_sh)
    open_frac = 1.0 - (reported_g / total_g)
    unc = None
    if land_diag:
        rms = {
            p: float(land_diag.get("resid_rms", {}).get(p, 0.0)) / 100.0
            for p in PARTIES
        }
        surp = {
            p: float(land_diag.get("surprise", {}).get(p, 0.0)) / 100.0
            for p in PARTIES
        }
        unc = uncertainty_halfwidth_pp(
            open_frac=open_frac,
            w_learn=float(land_diag.get("learn_weight", 0.0)),
            representativeness=float(land_diag.get("representativeness", 0.0)),
            resid_rms=rms,
            surprise=surp,
            frac_reported=1.0 - open_frac,
            n_units_open=len(unit_addrs) - len(reported_here),
            prior_unc_pp=prior_unc_pp,
        )
    out = {
        "frac_reported": round(reported_g / total_g, 4),
        "n_reported": len(reported_here),
        "nowcast": _pct_shares(nowcast),
        "naive": _pct_shares(naive),
        "mae_nowcast": round(mae(nowcast, truth_sh) * 100, 4),
        "mae_naive": round(mae(naive, truth_sh) * 100, 4),
    }
    if unc is not None:
        out["uncertainty"] = unc
    return out


def _build_neighbor_index(
    addrs: list[str],
    precincts: dict[str, dict],
    struktur: dict[str, np.ndarray],
    cfg: NowcastConfig,
) -> dict[str, list[str]]:
    by_art: dict[str, list[str]] = {"W": [], "B": []}
    for a in addrs:
        by_art.setdefault(precincts[a]["art"], []).append(a)

    idx: dict[str, list[str]] = {}
    for art, pool in by_art.items():
        valid = [a for a in pool if a in struktur]
        if len(valid) < 2:
            for a in pool:
                idx[a] = [x for x in pool if x != a][: cfg.k_neighbors]
            continue
        F = np.vstack([struktur[a] for a in valid])
        norms = np.linalg.norm(F, axis=1, keepdims=True) + EPS
        Fn = F / norms
        for start in range(0, len(valid), 200):
            end = min(start + 200, len(valid))
            sims = Fn[start:end] @ Fn.T
            for ii, i in enumerate(range(start, end)):
                row = sims[ii].copy()
                row[i] = -np.inf
                bez = precincts[valid[i]]["bezirk"]
                for j, b in enumerate(valid):
                    if precincts[b]["bezirk"] == bez:
                        row[j] += 0.35
                top = np.argpartition(-row, min(cfg.k_neighbors, len(valid) - 1))[
                    : cfg.k_neighbors
                ]
                idx[valid[i]] = [valid[j] for j in top]
        for a in pool:
            if a not in idx:
                idx[a] = [x for x in valid if x != a][: cfg.k_neighbors]
    return idx


def residual_rms(
    residuals: dict[str, dict[str, float]],
    surprise: dict[str, float],
    precincts: dict[str, dict],
    reported: set[str],
) -> dict[str, float]:
    """Vote-weighted RMS of precinct residuals around the global surprise."""
    if not reported:
        return {p: 0.0 for p in PARTIES}
    tw = sum(precincts[a]["gueltig"] for a in reported) + EPS
    out: dict[str, float] = {}
    for p in PARTIES:
        s = sum(
            precincts[a]["gueltig"] * (residuals[a][p] - surprise[p]) ** 2
            for a in reported
        )
        out[p] = float(np.sqrt(s / tw))
    return out


def uncertainty_halfwidth_pp(
    *,
    open_frac: float,
    w_learn: float,
    representativeness: float,
    resid_rms: dict[str, float],
    surprise: dict[str, float],
    frac_reported: float | None = None,
    n_units_open: int | None = None,
    prior_unc_pp: dict[str, float] | None = None,
) -> dict[str, float]:
    """Indicative half-band (≈ soft 80 %) in percentage points.

    Core: open_frac × forecast CI (party-specific). That alone shrinks
    steadily as votes come in. A modest selection/residual factor can
    *tighten* further when the sample looks informative, but cannot push
    the band above the shrinking open×prior envelope (avoids the old
    “stuck at prior for half the night” look).

    Replay also applies monotone_uncertainty as a safety net.
    """
    open_frac = float(np.clip(open_frac, 0.0, 1.0))
    frac = 1.0 - open_frac if frac_reported is None else float(
        np.clip(frac_reported, 0.0, 1.0)
    )
    sample_w = min(1.0, frac / 0.15)
    prior = prior_unc_pp or {p: 2.3 for p in PARTIES}
    out: dict[str, float] = {}
    for p in PARTIES:
        prior_sigma = max(0.005, float(prior.get(p, 2.3)) / 100.0)
        envelope = open_frac * prior_sigma
        if frac <= 0.0:
            out[p] = round(min(12.0, prior_sigma * 100.0), 2)
            continue
        # Informative night → slight tighten below the open×prior envelope
        tighten = 1.0 - 0.35 * float(w_learn) * float(representativeness) * sample_w
        half = envelope * max(0.55, tighten)
        raw_rms = float(resid_rms.get(p, 0.0))
        # Tiny residual bump, still capped by the envelope
        bump = open_frac * sample_w * raw_rms * 0.25
        if bump > 0:
            half = math.sqrt(half**2 + bump**2)
        if n_units_open is not None:
            idio = raw_rms / math.sqrt(max(1.0, float(n_units_open)))
            half = math.sqrt(half**2 + (open_frac * idio) ** 2)
        half = min(half, envelope)  # never wider than open × forecast CI
        out[p] = round(min(12.0, max(0.0, half) * 100.0), 2)
    return out


def monotone_uncertainty(
    prev: dict[str, float] | None,
    cur: dict[str, float] | None,
) -> dict[str, float] | None:
    """Party-wise running min so night bands never widen over steps."""
    if cur is None:
        return None
    if not prev:
        return {p: round(float(cur.get(p, 0.0)), 2) for p in PARTIES}
    return {
        p: round(min(float(prev.get(p, cur.get(p, 0.0))), float(cur.get(p, 0.0))), 2)
        for p in PARTIES
    }


NOWCAST_METHODS = (
    "prior",
    "naive",
    "global",
    "local",
    "full",
)
NOWCAST_METHOD_LABELS = {
    "prior": "Nur Ausgangslage",
    "naive": "Nur gemeldet (naiv)",
    "global": "Globaler Swing",
    "local": "Bezirk × Urne/Brief",
    "full": "Lokal + Nachbarn",
}


def nowcast_at(
    reported: set[str],
    precincts: dict[str, dict],
    prior_shares: dict[str, dict[str, float]],
    l1_shares: dict[str, dict[str, float]],
    struktur: dict[str, np.ndarray],
    cfg: NowcastConfig,
    neighbor_index: dict[str, list[str]] | None = None,
    *,
    method: str = "full",
    prior_unc_pp: dict[str, float] | None = None,
) -> tuple[dict[str, dict[str, float]], dict]:
    """Correct prior swing using surprise in reported precincts.

    Methods (open precincts only; reported always pinned to truth):
      prior  — keep prior swing, ignore night returns
      naive  — copy citywide reported aggregate onto every open precinct
      global — apply citywide surprise × learn_weight
      local  — Bezirk×Urne/Brief surprise mix (no structure neighbors)
      full   — local + structure neighbors (default product model)

    Returns (shares_by_addr, diagnostics).
    """
    if method not in NOWCAST_METHODS:
        raise ValueError(f"unknown nowcast method {method!r}")

    all_addrs = [a for a in precincts if a in prior_shares and a in l1_shares]
    truth_shares = {a: precincts[a]["shares"] for a in all_addrs}
    reported = {a for a in reported if a in all_addrs}

    total_g = sum(precincts[a]["gueltig"] for a in all_addrs) + EPS
    reported_g = sum(precincts[a]["gueltig"] for a in reported)
    frac = reported_g / total_g
    rep = composition_representativeness(reported, precincts, l1_shares, struktur)
    w_learn = learn_weight(frac, rep, cfg)

    def _diag(
        *,
        surprise: dict[str, float],
        rms: dict[str, float],
        w: float,
        r: float,
        f: float,
    ) -> dict:
        unc = uncertainty_halfwidth_pp(
            open_frac=1.0 - f,
            w_learn=w,
            representativeness=r,
            resid_rms=rms,
            surprise=surprise,
            frac_reported=f,
            n_units_open=len(all_addrs) - len(reported),
            prior_unc_pp=prior_unc_pp,
        )
        return {
            "method": method,
            "learn_weight": round(w, 4),
            "representativeness": round(r, 4),
            "frac_votes": round(f, 4),
            "surprise": {p: round(surprise[p] * 100, 3) for p in PARTIES},
            "resid_rms": {p: round(rms[p] * 100, 3) for p in PARTIES},
            "uncertainty": unc,
        }

    if not reported:
        surprise0 = {p: 0.0 for p in PARTIES}
        rms0 = {p: 0.0 for p in PARTIES}
        return {a: dict(prior_shares[a]) for a in all_addrs}, _diag(
            surprise=surprise0, rms=rms0, w=0.0, r=0.0, f=0.0
        )

    prior_r = aggregate_shares(precincts, prior_shares, list(reported))
    obs_r = aggregate_shares(precincts, truth_shares, list(reported))
    surprise = {p: obs_r[p] - prior_r[p] for p in PARTIES}
    residuals = {
        a: {p: truth_shares[a][p] - prior_shares[a][p] for p in PARTIES} for a in reported
    }
    rms = residual_rms(residuals, surprise, precincts, reported)

    if method == "prior":
        out = {
            a: dict(truth_shares[a]) if a in reported else dict(prior_shares[a])
            for a in all_addrs
        }
        return out, _diag(surprise=surprise, rms=rms, w=0.0, r=rep, f=frac)

    if method == "naive":
        out = {}
        for a in all_addrs:
            out[a] = dict(truth_shares[a]) if a in reported else dict(obs_r)
        return out, _diag(surprise=surprise, rms=rms, w=1.0, r=rep, f=frac)

    cell_surp: dict[tuple[str, str], dict[str, float]] = {}
    cell_w: dict[tuple[str, str], float] = {}
    if method in ("local", "full"):
        for a in reported:
            key = (precincts[a]["bezirk"], precincts[a]["art"])
            g = precincts[a]["gueltig"]
            cell_w[key] = cell_w.get(key, 0.0) + g
            if key not in cell_surp:
                cell_surp[key] = {p: 0.0 for p in PARTIES}
            for p in PARTIES:
                cell_surp[key][p] += residuals[a][p] * g
        for key, gw in cell_w.items():
            cell_surp[key] = {p: cell_surp[key][p] / (gw + EPS) for p in PARTIES}

    use_neighbors = method == "full" and neighbor_index
    mix_local = 0.0 if method == "global" else cfg.mix_local

    out = {}
    for a in all_addrs:
        if a in reported:
            out[a] = dict(truth_shares[a])
            continue
        if method == "global":
            local = surprise
        else:
            key = (precincts[a]["bezirk"], precincts[a]["art"])
            local = (
                cell_surp[key]
                if key in cell_surp and cell_w.get(key, 0) > 0
                else surprise
            )
            if use_neighbors and a in neighbor_index:
                neigh = [j for j in neighbor_index[a] if j in reported]
                if neigh:
                    tw = sum(precincts[j]["gueltig"] for j in neigh) + EPS
                    neigh_s = {
                        p: sum(
                            residuals[j][p] * precincts[j]["gueltig"] for j in neigh
                        )
                        / tw
                        for p in PARTIES
                    }
                    local = {p: 0.5 * local[p] + 0.5 * neigh_s[p] for p in PARTIES}
        corr = {
            p: w_learn * (mix_local * local[p] + (1.0 - mix_local) * surprise[p])
            for p in PARTIES
        }
        out[a] = _renorm({p: prior_shares[a][p] + corr[p] for p in PARTIES})

    return out, _diag(surprise=surprise, rms=rms, w=w_learn, r=rep, f=frac)


def load_reporting_times_2023() -> dict[str, datetime]:
    """Per-Wahlbezirk Datum/Zeit from AfS Datenexport _W_ (frozen end file).

    Note: in the published end-state these often look like post-election
    finalization times (from 13.02.), not live election-night entry. Still the
    official per-precinct chronology available without night scrapes.
    """
    path = RAW / "Datenexport_AGH2023_Zweitstimme_W_BE.csv"
    if not path.exists():
        return {}
    out: dict[str, datetime] = {}
    with path.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f, delimiter=";"):
            addr = (row.get("Adresse") or "").strip()
            if not addr:
                continue
            try:
                out[addr] = datetime.strptime(
                    (row.get("Datum") or "") + (row.get("Zeit") or ""),
                    "%y.%m.%d%H:%M:%S",
                )
            except ValueError:
                continue
    return out


# AGH 2023 election evening (approx.) for simulated Meldeordnungen without AfS times
_SIM_NIGHT_START = datetime(2023, 2, 12, 18, 0, 0)
_SIM_NIGHT_HOURS = 8.0


def clock_for_step(
    *,
    order: list[str],
    n_reported: int,
    frac: float,
    reporting_times: dict[str, datetime] | None,
    scenario: str,
) -> tuple[str | None, str]:
    """Return (clock 'YYYY-MM-DD HH:MM:SS', source 'afs'|'simulated'|'')."""
    times = reporting_times or {}
    if scenario == "actual_times" and times and n_reported > 0:
        last = order[min(n_reported, len(order)) - 1]
        t = times.get(last)
        if t is None:
            # fall back: max time among reported prefix
            known = [times[a] for a in order[:n_reported] if a in times]
            t = max(known) if known else None
        if t is not None:
            return t.strftime("%Y-%m-%d %H:%M:%S"), "afs"
    # Simulated / missing AfS times: map Auszählungsanteil onto election night
    t = _SIM_NIGHT_START + timedelta(
        hours=_SIM_NIGHT_HOURS * float(np.clip(frac, 0.0, 1.0))
    )
    return t.strftime("%Y-%m-%d %H:%M:%S"), "simulated"


def clock_when_addrs_complete(
    *,
    addrs: list[str] | set[str],
    order: list[str],
    reporting_times: dict[str, datetime] | None,
    scenario: str,
    land_frac: float,
    step_clock: str | None = None,
    step_clock_source: str | None = None,
) -> tuple[str | None, str]:
    """Uhrzeit, zu der der letzte Wahlbezirk eines Gebiets gemeldet war."""
    want = set(addrs)
    last = None
    for a in order:
        if a in want:
            last = a
    if last is None:
        return step_clock, (step_clock_source or "")
    try:
        pos = order.index(last) + 1
        frac = pos / max(len(order), 1)
    except ValueError:
        pos = 0
        frac = land_frac
    times = reporting_times or {}
    if scenario == "actual_times" and times:
        t = times.get(last)
        if t is None:
            known = [times[a] for a in want if a in times]
            t = max(known) if known else None
        if t is not None:
            return t.strftime("%Y-%m-%d %H:%M:%S"), "afs"
    return clock_for_step(
        order=order,
        n_reported=pos,
        frac=frac,
        reporting_times=None,
        scenario="simulated",
    )


def simulate_order(
    precincts: dict[str, dict],
    scenario: str,
    rng: np.random.Generator,
    reporting_times: dict[str, datetime] | None = None,
) -> list[str]:
    addrs = list(precincts.keys())
    if scenario == "actual_times":
        times = reporting_times or {}
        missing = [a for a in addrs if a not in times]
        known = [a for a in addrs if a in times]
        known.sort(key=lambda a: (times[a], a))
        # precincts without a timestamp after those with times (stable)
        missing.sort()
        return known + missing
    if scenario == "random":
        rng.shuffle(addrs)
        return addrs
    if scenario == "urne_first":
        w = [a for a in addrs if precincts[a]["art"] == "W"]
        b = [a for a in addrs if precincts[a]["art"] != "W"]
        rng.shuffle(w)
        rng.shuffle(b)
        return w + b
    if scenario == "small_first":
        return sorted(addrs, key=lambda a: (precincts[a]["wber"], a))
    if scenario == "large_first":
        return sorted(addrs, key=lambda a: (-precincts[a]["wber"], a))
    if scenario == "green_first":
        return sorted(addrs, key=lambda a: (-precincts[a]["shares"]["gruene"], a))
    if scenario == "cdu_first":
        return sorted(addrs, key=lambda a: (-precincts[a]["shares"]["cdu"], a))
    raise ValueError(f"unknown scenario {scenario}")


def mae(a: dict[str, float], b: dict[str, float]) -> float:
    return float(np.mean([abs(a[p] - b[p]) for p in PARTIES]))


def _leader(shares: dict[str, float]) -> str:
    """Largest among main parties (shares as fractions or percent — same argmax)."""
    return max(MAIN_PARTIES, key=lambda p: shares.get(p, 0.0))


def _above_hurdle(shares: dict[str, float]) -> dict[str, bool]:
    """shares as fractions."""
    return {p: shares.get(p, 0.0) >= HURDLE for p in MAIN_PARTIES}


def _phi(x: float) -> float:
    """Standard normal CDF."""
    return 0.5 * (1.0 + math.erf(x / math.sqrt(2.0)))


def inflate_unc_open_prior(
    unc: dict[str, float] | None,
    frac_reported: float,
    *,
    floor_pp: float = WK_OPEN_PRIOR_FLOOR_PP,
) -> dict[str, float]:
    """Add irreducible prior-race floor while a unit is still largely open."""
    base = dict(unc or {})
    open_w = 1.0 - float(np.clip(frac_reported, 0.0, 1.0))
    floor = float(floor_pp) * open_w
    if floor <= 1e-9:
        return {p: round(float(base.get(p, 0.0)), 2) for p in PARTIES}
    out: dict[str, float] = {}
    for p in PARTIES:
        out[p] = round(math.sqrt(float(base.get(p, 0.0)) ** 2 + floor**2), 2)
    return out


def p_lead_from_margin(
    margin_pp: float,
    unc_pp: dict[str, float],
    top2: list[str],
    *,
    complete: bool,
) -> float:
    if complete:
        return 1.0
    sigma = math.sqrt(
        (float(unc_pp.get(top2[0], 0.0)) / 1.28) ** 2
        + (float(unc_pp.get(top2[1], 0.0)) / 1.28) ** 2
    )
    if sigma <= 1e-9:
        return 1.0 if margin_pp > 0 else 0.5
    return float(_phi(margin_pp / sigma))


def _hare_generic(votes: dict[str, float], seats: int) -> dict[str, int]:
    """Hare/Niemeyer over arbitrary keys (e.g. Bezirke)."""
    keys = [k for k, v in votes.items() if v > 0]
    if not keys or seats <= 0:
        return {k: 0 for k in votes}
    v = np.array([votes[k] for k in keys], dtype=float)
    quota = float(v.sum()) / seats
    exact = v / quota
    base = np.floor(exact).astype(int)
    rem = seats - int(base.sum())
    order = np.argsort(-(exact - base))
    for i in range(rem):
        base[order[i]] += 1
    out = {k: 0 for k in votes}
    for k, n in zip(keys, base):
        out[k] = int(n)
    return out


def hare_niemeyer(votes: dict[str, float], seats: int) -> dict[str, int]:
    parties = [p for p in MAIN_PARTIES if votes.get(p, 0.0) > 0]
    if not parties or seats <= 0:
        return {p: 0 for p in MAIN_PARTIES}
    v = np.array([votes[p] for p in parties], dtype=float)
    total = float(v.sum())
    if total <= 0:
        return {p: 0 for p in MAIN_PARTIES}
    quota = total / seats
    exact = v / quota
    base = np.floor(exact).astype(int)
    rem = seats - int(base.sum())
    order = np.argsort(-(exact - base))
    for i in range(rem):
        base[order[i]] += 1
    out = {p: 0 for p in MAIN_PARTIES}
    for p, n in zip(parties, base):
        out[p] = int(n)
    return out


def list_seats_from_shares(
    shares: dict[str, float], seats: int = AGH_BASE_SEATS
) -> dict[str, int]:
    """Hare/Niemeyer among parties clearing 5% (no Direkt/Grundmandat)."""
    eligible = {p: shares[p] for p in MAIN_PARTIES if shares.get(p, 0.0) >= HURDLE}
    return hare_niemeyer(eligible, seats)


def seat_mae(a: dict[str, int], b: dict[str, int]) -> float:
    return float(np.mean([abs(a.get(p, 0) - b.get(p, 0)) for p in MAIN_PARTIES]))


def allocate_be(
    votes: dict[str, float],
    directs: dict[str, int],
    base: int = AGH_BASE_SEATS,
    *,
    directs_by_bez: dict[str, dict] | None = None,
    bez_votes: dict[str, dict] | None = None,
    bezirk_parties: set[str] | list[str] | None = None,
) -> dict:
    """Berlin AGH: same allocator as parliament_size_sim (Bezirk overhang)."""
    from parliament_size_sim import allocate_be as _alloc

    res = _alloc(
        votes,
        directs,
        base,
        directs_by_bez=directs_by_bez,
        bez_votes=bez_votes,
        bezirk_parties=bezirk_parties if bezirk_parties is not None else set(
            BE_BEZIRKSLISTE_PARTIES
        ),
    )
    seats = {p: 0 for p in MAIN_PARTIES}
    seats.update(res.get("seats") or {})
    return {
        "size": res["size"],
        "seats": seats,
        "alloc": res.get("alloc") or dict(seats),
        "total_oh": res.get("total_oh", 0),
        "incomplete": res.get("incomplete", False),
        "adv": res.get("adv", 0),
    }


def load_erst_winners_2023() -> dict[str, str]:
    """Citywide WK id → Erststimme winner party code (AGH 2023)."""
    rows = _read_xlsx_sheet(RAW / "DL_BE_AGHBVV2023.xlsx", "AGH_W1")
    votes: dict[tuple[str, int], dict[str, float]] = {}
    for r in rows:
        bez = str(r.get("Bezirksnummer") or "").zfill(2)
        local = int(
            _f(
                r.get("Abgeordnetenhauswahlkreis")
                or r.get("Abgeordneten-hauswahlkreis")
            )
        )
        key = (bez, local)
        cell = votes.setdefault(key, {p: 0.0 for p in MAIN_PARTIES})
        for col, code in XLSX_PARTY_MAP.items():
            if code == "others":
                continue
            cell[code] += _f(r.get(col))
    pairs = sorted(votes)
    out: dict[str, str] = {}
    for i, key in enumerate(pairs, start=1):
        out[str(i)] = max(MAIN_PARTIES, key=lambda p: votes[key][p])
    return out


def load_direkt_roster_placeholders() -> dict[str, dict[str, dict]]:
    """Placeholder Direktkandidaten for the AGH2023 replay UI.

    Replay evaluates 2023 Erst winners; we do not ship 2023 person names, and
    2026 nominees would be anachronistic — so every cell is a Platzhalter.
    """
    by_wkr: dict[str, dict[str, dict]] = {str(i): {} for i in range(1, 79)}
    for wid, parties in by_wkr.items():
        for p in MAIN_PARTIES:
            parties[p] = {
                "name": f"{PARTY_LABELS[p]} · WK {wid} · Platzhalter",
                "is_placeholder": True,
                "source": "",
            }
    return by_wkr


def load_listen_roster_placeholders() -> dict[str, dict]:
    """Listen structure from 2026 CSV, but all display names anonymized.

    Keeps list_type / positions / Direkt-Doppel (wkr) for Einzug-UI layout;
    names are Platzhalter because the replay is AGH2023.
    """
    path = REPO / "berlin" / "candidates" / "listenkandidaten_2026.csv"
    out: dict[str, dict] = {}
    if not path.exists():
        return out
    with path.open(newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            party = (r.get("party") or "").strip().lower()
            if party not in MAIN_PARTIES:
                continue
            try:
                pos = int(r.get("list_pos") or 0)
            except (TypeError, ValueError):
                continue
            if pos <= 0:
                continue
            lt = (r.get("list_type") or "").strip().lower() or "landes"
            label = PARTY_LABELS.get(party, party.upper())
            entry: dict = {
                "pos": pos,
                "name": f"{label} · Listenplatz {pos}",
                "ph": True,
            }
            wkd = (r.get("wkr_direct") or "").strip()
            if wkd:
                try:
                    entry["wkr"] = int(float(wkd))
                except ValueError:
                    pass
            slot = out.setdefault(
                party, {"list_type": lt, "landes": [], "bezirk": {}}
            )
            bez = (r.get("bezirk") or "").strip()
            if lt == "bezirk" and bez:
                slot["bezirk"].setdefault(bez.zfill(2), []).append(entry)
            else:
                slot["landes"].append(entry)
    for slot in out.values():
        slot["landes"].sort(key=lambda e: e["pos"])
        for entries in slot["bezirk"].values():
            entries.sort(key=lambda e: e["pos"])
    return out


def night_entry_mc(
    nc_land_pct: dict[str, float],
    unc_pp: dict[str, float],
    by_bez_pct: dict[str, dict[str, float]],
    wkr_leaders: dict[str, str],
    wkr_bez: dict[str, str],
    rng: np.random.Generator,
    n_draws: int = 48,
) -> dict:
    """Monte Carlo über das Unsicherheitsband: Größe, Sitze, Listensitze.

    Vereinfachungen: Direktmandate fix beim aktuellen Nowcast-Führer;
    Bezirks-Anteile werden proportional zum Landes-Draw mitverschoben;
    Bezirks-Suballokation via Hare/Niemeyer ohne Direktmandats-Garantie
    (Listensitze werden bei 0 gekappt).
    Quantile [p10, p50, p90] ≈ 80-%-Intervall.
    """
    directs = {p: 0 for p in MAIN_PARTIES}
    directs_bez: dict[tuple[str, str], int] = {}
    for uid, lead in wkr_leaders.items():
        if lead not in directs:
            continue
        directs[lead] += 1
        key = (lead, wkr_bez.get(uid, ""))
        directs_bez[key] = directs_bez.get(key, 0) + 1
    directs_by_bez: dict[str, dict[str, int]] = {p: {} for p in MAIN_PARTIES}
    for (p, b), n in directs_bez.items():
        if b:
            directs_by_bez.setdefault(p, {})[str(b).zfill(2)] = n

    bez_ids = sorted(by_bez_pct)
    sizes: list[int] = []
    seats_acc: dict[str, list[int]] = {p: [] for p in MAIN_PARTIES}
    list_land_acc: dict[str, list[int]] = {p: [] for p in MAIN_PARTIES}
    list_bez_acc: dict[str, dict[str, list[int]]] = {
        p: {b: [] for b in bez_ids} for p in BE_BEZIRKSLISTE_PARTIES
    }
    for _ in range(n_draws):
        draw = {}
        for p in PARTIES:
            sd = float(unc_pp.get(p, 0.0)) / 1.28  # Band ≈ 80 % ⇒ z ≈ 1.28
            draw[p] = max(0.0, float(nc_land_pct.get(p, 0.0)) + rng.normal(0.0, sd))
        tot = sum(draw.values()) or 1.0
        frac = {p: draw[p] / tot for p in PARTIES}
        bez_votes = {}
        for p in MAIN_PARTIES:
            scale = draw[p] / max(float(nc_land_pct.get(p, 0.0)), EPS)
            bez_votes[p] = {
                b: max(0.0, by_bez_pct[b].get(p, 0.0) * scale) for b in bez_ids
            }
        alloc = allocate_be(
            frac,
            directs,
            directs_by_bez=directs_by_bez,
            bez_votes=bez_votes,
            bezirk_parties=set(BE_BEZIRKSLISTE_PARTIES),
        )
        sizes.append(int(alloc["size"]))
        for p in MAIN_PARTIES:
            s_p = int(alloc["seats"].get(p, 0))
            seats_acc[p].append(s_p)
            list_land_acc[p].append(max(0, s_p - directs.get(p, 0)))
        for p in BE_BEZIRKSLISTE_PARTIES:
            s_p = int(alloc["seats"].get(p, 0))
            if s_p <= 0:
                for b in bez_ids:
                    list_bez_acc[p][b].append(0)
                continue
            scale = draw[p] / max(float(nc_land_pct.get(p, 0.0)), EPS)
            votes = {
                b: max(0.0, by_bez_pct[b].get(p, 0.0) * scale) for b in bez_ids
            }
            bseats = _hare_generic(votes, s_p)
            for b in bez_ids:
                list_bez_acc[p][b].append(
                    max(0, bseats.get(b, 0) - directs_bez.get((p, b), 0))
                )

    def q(vals: list[int]) -> list[int]:
        arr = np.asarray(vals)
        return [
            int(np.percentile(arr, 10)),
            int(np.percentile(arr, 50)),
            int(np.percentile(arr, 90)),
        ]

    list_seats: dict[str, object] = {}
    for p in MAIN_PARTIES:
        if p in BE_BEZIRKSLISTE_PARTIES:
            list_seats[p] = {b: q(list_bez_acc[p][b]) for b in bez_ids}
        else:
            list_seats[p] = q(list_land_acc[p])
    return {
        "n_draws": n_draws,
        "size": q(sizes),
        "seats": {p: q(seats_acc[p]) for p in MAIN_PARTIES},
        "directs": directs,
        "list_seats": list_seats,
    }


# Scenario-config party codes (gru/lin) → our codes (gruene/linke)
_SCENARIO_PARTY_ALIAS = {
    "gru": "gruene",
    "gruene": "gruene",
    "lin": "linke",
    "linke": "linke",
    "spd": "spd",
    "cdu": "cdu",
    "afd": "afd",
    "fdp": "fdp",
    "bsw": "bsw",
    "others": "others",
}


def _scenario_party(p: str) -> str:
    return _SCENARIO_PARTY_ALIAS.get((p or "").strip().lower(), (p or "").strip().lower())


def _coalition_has_majority(
    shares: dict[str, float],
    parties: list[str],
    *,
    hurdle: float = 0.05,
) -> bool:
    """Same rule as state-model/compute_scenarios.R (Zweit after hurdle)."""
    parl = {p: v for p, v in shares.items() if p != "others"}
    for p in parties:
        if parl.get(p, 0.0) < hurdle:
            return False
    above = sum(v for v in parl.values() if v >= hurdle)
    if above <= EPS:
        return False
    return sum(parl.get(p, 0.0) for p in parties) / above > 0.5


def _eval_scenario(shares: dict[str, float], defn: dict) -> bool:
    cat = defn["category"]
    if cat == "largest_party":
        p = defn["party"]
        parl = {k: v for k, v in shares.items() if k != "others"}
        if p not in parl:
            return False
        return parl[p] >= max(parl.values()) - EPS
    if cat == "above_hurdle":
        return shares.get(defn["party"], 0.0) >= float(defn.get("hurdle", 0.05))
    if cat == "coalition":
        parties = defn["parties"]
        if not _coalition_has_majority(
            shares, parties, hurdle=float(defn.get("hurdle", 0.05))
        ):
            return False
        lead = defn.get("lead")
        if not lead:
            return True
        return shares.get(lead, 0.0) >= max(shares.get(p, 0.0) for p in parties) - EPS
    return False


def load_be_scenario_defs(*, hurdle: float = 0.05) -> list[dict]:
    """Political scenarios for Berlin, aligned with Szenario-Wahrscheinlichkeiten."""
    path = REPO / "data" / "state_forecast_scenarios.json"
    if not path.exists():
        return []
    cfg = json.loads(path.read_text(encoding="utf-8"))
    defs: list[dict] = []
    known = set(MAIN_PARTIES)  # no BSW in AGH2023 party bag

    for p in MAIN_PARTIES:
        defs.append(
            {
                "id": f"largest_party_{'gru' if p == 'gruene' else ('lin' if p == 'linke' else p)}",
                "category": "largest_party",
                "label_de": f"{PARTY_LABELS[p]} stärkste Kraft",
                "party": p,
                "hurdle": hurdle,
            }
        )

    for raw in cfg.get("above_hurdle_parties_by_state", {}).get("BE", ["fdp"]):
        p = _scenario_party(raw)
        if p not in known:
            continue
        defs.append(
            {
                "id": f"above_hurdle_{raw}",
                "category": "above_hurdle",
                "label_de": f"{PARTY_LABELS[p]} über 5%-Hürde",
                "party": p,
                "hurdle": hurdle,
            }
        )

    coalitions = list(cfg.get("coalitions") or [])
    coalitions.extend(cfg.get("coalitions_by_state", {}).get("BE") or [])
    excluded = set(cfg.get("exclude_scenario_ids_by_state", {}).get("BE") or [])
    seen: set[str] = set()
    for coal in coalitions:
        cid = coal.get("id") or ""
        if not cid or cid in excluded or cid in seen:
            continue
        parties = [_scenario_party(x) for x in (coal.get("parties") or [])]
        if any(p not in known for p in parties):
            continue  # skip BSW-coalitions in 2023 replay
        seen.add(cid)
        lead_raw = coal.get("lead")
        lead = _scenario_party(lead_raw) if lead_raw else None
        label = (coal.get("label_de") or cid).replace("CDU/CSU", "CDU")
        defs.append(
            {
                "id": cid,
                "category": "coalition",
                "label_de": label,
                "parties": parties,
                "lead": lead,
                "hurdle": hurdle,
            }
        )
    return defs


def night_scenario_probs(
    nc_land_pct: dict[str, float],
    unc_pp: dict[str, float],
    truth_land_pct: dict[str, float],
    prior_land_pct: dict[str, float] | None = None,
    rng: np.random.Generator | None = None,
    *,
    n_draws: int = 48,
    scenario_defs: list[dict] | None = None,
    p_start_by_id: dict[str, float] | None = None,
) -> dict:
    """P(Szenario) aus Nowcast-Band — Call bei P≥50 % vs. Wahrheit.

    Inputs are percentages (0–100). ``p_start`` = Wkt vor Auszählungsbeginn
    (Step-0-MC), falls ``p_start_by_id`` gesetzt; sonst = aktuelles ``p``.
    ``prior_land_pct`` ist ungenutzt (API-Kompatibilität).
    """
    del prior_land_pct  # früher: deterministische 0/100-Ausgangslage
    if rng is None:
        rng = np.random.default_rng(0)
    defs = scenario_defs if scenario_defs is not None else load_be_scenario_defs()
    if not defs:
        return {
            "n_draws": n_draws,
            "call_threshold": 0.5,
            "items": [],
            "n_ok": 0,
            "n_total": 0,
        }

    def as_frac(pct: dict[str, float]) -> dict[str, float]:
        return {p: float(pct.get(p, 0.0)) / 100.0 for p in PARTIES}

    truth_frac = as_frac(truth_land_pct)

    hits = {d["id"]: 0 for d in defs}
    for _ in range(n_draws):
        draw = {
            p: max(
                0.0,
                float(nc_land_pct.get(p, 0.0))
                + rng.normal(0.0, float(unc_pp.get(p, 0.0)) / 1.28),
            )
            for p in PARTIES
        }
        tot = sum(draw.values()) or 1.0
        frac = {p: draw[p] / tot for p in PARTIES}
        for d in defs:
            if _eval_scenario(frac, d):
                hits[d["id"]] += 1

    items = []
    n_ok = 0
    for d in defs:
        p_hat = hits[d["id"]] / float(n_draws)
        truth = bool(_eval_scenario(truth_frac, d))
        call = p_hat >= 0.5
        correct = call == truth
        if correct:
            n_ok += 1
        p_now = round(p_hat * 100.0, 1)
        if p_start_by_id is not None and d["id"] in p_start_by_id:
            p_start = round(float(p_start_by_id[d["id"]]), 1)
        else:
            p_start = p_now
        items.append(
            {
                "id": d["id"],
                "category": d["category"],
                "label_de": d["label_de"],
                "p": p_now,
                "p_start": p_start,
                "truth": truth,
                "call": call,
                "correct": correct,
            }
        )
    items.sort(key=lambda x: (-x["p"], x["label_de"]))
    return {
        "n_draws": n_draws,
        "call_threshold": 0.5,
        "items": items,
        "n_ok": n_ok,
        "n_total": len(items),
    }


def attach_scenario_p_start(steps: list[dict]) -> None:
    """Annotate each step's scenario items with p_start from step 0 (pre-count)."""
    if not steps:
        return
    start_items = (steps[0].get("scenario_probs") or {}).get("items") or []
    start_p = {it["id"]: it["p"] for it in start_items if "id" in it and "p" in it}
    for s in steps:
        sp = s.get("scenario_probs") or {}
        for it in sp.get("items") or []:
            sid = it.get("id")
            if sid in start_p:
                it["p_start"] = start_p[sid]
            else:
                it.setdefault("p_start", it.get("p"))
            it.pop("p_prior", None)


def _directs_from_leaders(leaders: dict[str, str]) -> dict[str, int]:
    d = {p: 0 for p in MAIN_PARTIES}
    for lead in leaders.values():
        if lead in d:
            d[lead] += 1
    return d


def _wkr_leaders_from_step(
    by_wkr: dict[str, dict],
    *,
    pred_key: str,
) -> dict[str, str]:
    return {
        uid: _leader(_as_frac(step[pred_key]))
        for uid, step in by_wkr.items()
        if pred_key in step
    }


def eval_night_institutions(
    pred_land: dict[str, float],
    truth_land: dict[str, float],
    by_wkr: dict[str, dict],
    *,
    pred_key: str,
    erst_winners_truth: dict[str, str],
    geo_static_wkr: dict[str, dict] | None = None,
    erst_known: dict[str, str] | None = None,
) -> dict:
    """Direktmandate (Erststimme), BE-Parlamentsgröße/Sitze, Parteieneinzug.

    erst_known: echte Erst-Sieger für bereits voll ausgezählte WKs — dort
    liegt das Ergebnis vor. Damit konvergiert die Vorhersage bei 100 % exakt
    auf die Wahrheit.
    """
    if pred_key == "prior" and geo_static_wkr:
        src = {uid: {"prior": st["prior"]} for uid, st in geo_static_wkr.items()}
        key = "prior"
    else:
        src = by_wkr
        key = pred_key
    pred_leaders = _wkr_leaders_from_step(src, pred_key=key)
    if erst_known:
        pred_leaders = {**pred_leaders, **erst_known}
    pred_directs = _directs_from_leaders(pred_leaders)
    truth_directs = _directs_from_leaders(erst_winners_truth)
    n_hit = 0
    n_tot = 0
    for uid, truth_w in erst_winners_truth.items():
        if uid not in pred_leaders:
            continue
        n_tot += 1
        if pred_leaders[uid] == truth_w:
            n_hit += 1
    alloc_pred = allocate_be(pred_land, pred_directs)
    alloc_truth = allocate_be(truth_land, truth_directs)
    entry_pred = {
        p: pred_land.get(p, 0.0) >= HURDLE or pred_directs.get(p, 0) >= 1
        for p in MAIN_PARTIES
    }
    entry_truth = {
        p: truth_land.get(p, 0.0) >= HURDLE or truth_directs.get(p, 0) >= 1
        for p in MAIN_PARTIES
    }
    entry_ok = sum(1 for p in MAIN_PARTIES if entry_pred[p] == entry_truth[p])
    return {
        "direkt": {
            "n_hit": n_hit,
            "n_total": n_tot,
            "hit_rate": round(n_hit / n_tot, 4) if n_tot else 0.0,
            "directs_pred": pred_directs,
            "directs_truth": truth_directs,
            "note": "Direktmandat = Erststimmen-Nowcast je WK vs. Erststimme-Sieger",
        },
        "parliament": {
            "size_pred": alloc_pred["size"],
            "size_truth": alloc_truth["size"],
            "size_err": alloc_pred["size"] - alloc_truth["size"],
            "seats_pred": alloc_pred["seats"],
            "seats_truth": alloc_truth["seats"],
            "seat_mae": round(seat_mae(alloc_pred["seats"], alloc_truth["seats"]), 4),
            "total_oh_pred": alloc_pred["total_oh"],
            "total_oh_truth": alloc_truth["total_oh"],
        },
        "entry": {
            "pred": entry_pred,
            "truth": entry_truth,
            "n_ok": entry_ok,
            "n_total": len(MAIN_PARTIES),
            "note": "Einzug Liste: ≥5 % oder ≥1 Direktmandat (Grundmandat)",
        },
    }


def _as_frac(shares: dict[str, float]) -> dict[str, float]:
    """Accept percent (sum~100) or fraction (sum~1)."""
    s = sum(shares.get(p, 0.0) for p in PARTIES)
    if s > 1.5:
        return {p: shares.get(p, 0.0) / 100.0 for p in PARTIES}
    return {p: shares.get(p, 0.0) for p in PARTIES}


def eval_list_forecast(
    pred_pct: dict[str, float],
    truth_pct: dict[str, float],
) -> dict:
    pred = _as_frac(pred_pct)
    truth = _as_frac(truth_pct)
    pred_h = _above_hurdle(pred)
    truth_h = _above_hurdle(truth)
    hurdle_ok = sum(1 for p in MAIN_PARTIES if pred_h[p] == truth_h[p])
    pred_seats = list_seats_from_shares(pred)
    truth_seats = list_seats_from_shares(truth)
    return {
        "top1_ok": _leader(pred) == _leader(truth),
        "top1_pred": _leader(pred),
        "top1_truth": _leader(truth),
        "hurdle_acc": round(hurdle_ok / len(MAIN_PARTIES), 4),
        "hurdle_ok": {p: pred_h[p] == truth_h[p] for p in MAIN_PARTIES},
        "above5_pred": {p: pred_h[p] for p in MAIN_PARTIES},
        "above5_truth": {p: truth_h[p] for p in MAIN_PARTIES},
        "seats_pred": pred_seats,
        "seats_truth": truth_seats,
        "seat_mae": round(seat_mae(pred_seats, truth_seats), 4),
    }


def eval_unit_winners(
    by_unit: dict[str, dict],
    static: dict[str, dict],
    *,
    pred_key: str = "nowcast",
) -> dict:
    """Share of units where predicted leader matches truth leader."""
    hits = 0
    n = 0
    for uid, step_u in by_unit.items():
        st = static.get(uid)
        if not st:
            continue
        n += 1
        pred_lead = _leader(_as_frac(step_u[pred_key]))
        truth_lead = _leader(_as_frac(st["truth"]))
        if pred_lead == truth_lead:
            hits += 1
    return {
        "hit_rate": round(hits / n, 4) if n else 0.0,
        "n_hit": hits,
        "n_total": n,
    }


def build_step_eval(
    nowcast_land: dict[str, float],
    naive_land: dict[str, float],
    prior_land: dict[str, float],
    truth_land: dict[str, float],
    by_bezirk: dict[str, dict],
    by_wkr: dict[str, dict],
    geo_static: dict,
    erst_winners_truth: dict[str, str] | None = None,
) -> dict:
    """List + WK/Bezirk + Direkt/Parlament/Einzug beyond share MAE."""
    erst = erst_winners_truth or {}
    out = {
        "list": {
            "nowcast": eval_list_forecast(_pct_shares(nowcast_land), _pct_shares(truth_land)),
            "naive": eval_list_forecast(_pct_shares(naive_land), _pct_shares(truth_land)),
            "prior": eval_list_forecast(_pct_shares(prior_land), _pct_shares(truth_land)),
        },
        "bezirk": {
            "nowcast": eval_unit_winners(
                by_bezirk, geo_static["bezirk"], pred_key="nowcast"
            ),
            "naive": eval_unit_winners(by_bezirk, geo_static["bezirk"], pred_key="naive"),
        },
        "wkr": {
            "nowcast": eval_unit_winners(by_wkr, geo_static["wkr"], pred_key="nowcast"),
            "naive": eval_unit_winners(by_wkr, geo_static["wkr"], pred_key="naive"),
        },
        "note": (
            "Liste: 5-%-Hürde + Top-1 + einfache Hare-Sitze (Zweitstimme). "
            "Direkt/Parlament: Erststimmen-Nowcast je WK + BE-Formel "
            "(Grundmandat, Überhang/Ausgleich). "
            "Bezirk: Zweit-Anteile (Bezirkslisten)."
        ),
    }
    if erst:
        # Voll ausgezählte WKs: echter Erst-Sieger bekannt → Proxy ersetzen.
        erst_known = {
            uid: erst[uid]
            for uid, r in by_wkr.items()
            if uid in erst and float(r.get("frac_reported", 0.0)) >= 0.999
        }
        out["institutions"] = {
            "nowcast": eval_night_institutions(
                nowcast_land,
                truth_land,
                by_wkr,
                pred_key="nowcast",
                erst_winners_truth=erst,
                geo_static_wkr=geo_static.get("wkr"),
                erst_known=erst_known,
            ),
            "naive": eval_night_institutions(
                naive_land,
                truth_land,
                by_wkr,
                pred_key="naive",
                erst_winners_truth=erst,
                geo_static_wkr=geo_static.get("wkr"),
                erst_known=erst_known,
            ),
            "prior": eval_night_institutions(
                prior_land,
                truth_land,
                by_wkr,
                pred_key="prior",
                erst_winners_truth=erst,
                geo_static_wkr=geo_static.get("wkr"),
            ),
        }
    return out


def run_replay(
    precincts: dict[str, dict],
    l1: dict[str, dict],
    struktur: dict[str, np.ndarray],
    prior_target: dict[str, float],
    prior_meta: dict | None = None,
    reporting_times: dict[str, datetime] | None = None,
    *,
    scenario: str = "urne_first",
    n_steps: int = 40,
    cfg: NowcastConfig | None = None,
    seed: int = RNG_SEED,
) -> dict:
    cfg = cfg or NowcastConfig()
    rng = np.random.default_rng(seed)
    addrs = sorted(set(precincts) & set(l1))
    precincts = {a: precincts[a] for a in addrs}
    l1 = {a: l1[a] for a in addrs}
    l1_shares = {a: l1[a]["shares"] for a in addrs}

    truth_land = aggregate_shares(precincts, {a: precincts[a]["shares"] for a in addrs})
    prior_shares = build_priors(precincts, l1, prior_target=prior_target)
    prior_land = aggregate_shares(precincts, prior_shares)
    prior_unc_pp = prior_uncertainty_pp(prior_meta)
    scope_addrs = _addrs_by_scope(precincts)
    geo_units = _geo_catalog(precincts)

    # Erststimme track: truth AGH_W1, L1 from 2016 Erst (Adresse-Overlap) +
    # OLS prior from Zweit-Prior (district_model_coefs).
    erst_raw = load_truth_erst_2023()
    erst_l1, erst_l1_meta = load_baseline_erst_2016(l1)
    beta_erst, _sigma_erst = load_erst_ols_coef()
    erst_prior = build_erst_priors(prior_shares, erst_l1, beta_erst)
    precincts_erst: dict[str, dict] = {}
    for a in addrs:
        if a not in erst_raw:
            continue
        er = erst_raw[a]
        precincts_erst[a] = {
            **{k: precincts[a][k] for k in ("bezirk", "bezirk_name", "art", "wkr", "wkr_local")},
            "gueltig": er["gueltig"],
            "shares": er["shares"],
        }
    # Align to common address set (Zweit ∩ Erst ∩ L1)
    addrs = sorted(set(addrs) & set(precincts_erst) & set(erst_prior) & set(erst_l1))
    precincts = {a: precincts[a] for a in addrs}
    precincts_erst = {a: precincts_erst[a] for a in addrs}
    l1 = {a: l1[a] for a in addrs}
    l1_shares = {a: l1[a]["shares"] for a in addrs}
    prior_shares = {a: prior_shares[a] for a in addrs}
    erst_prior = {a: erst_prior[a] for a in addrs}
    erst_l1_shares = {a: erst_l1[a]["shares"] for a in addrs}
    scope_addrs = _addrs_by_scope(precincts)
    geo_units = _geo_catalog(precincts)
    truth_land = aggregate_shares(precincts, {a: precincts[a]["shares"] for a in addrs})
    prior_land = aggregate_shares(precincts, prior_shares)

    geo_static = {
        "bezirk": {
            uid: _region_truth_prior(ua, precincts, prior_shares)
            for uid, ua in sorted(scope_addrs["bezirk"].items())
        },
        # WK-Ansicht / Direkt = Erststimme
        "wkr": {
            uid: _region_truth_prior(ua, precincts_erst, erst_prior)
            for uid, ua in sorted(
                scope_addrs["wkr"].items(), key=lambda kv: int(kv[0])
            )
        },
    }
    erst_winners = load_erst_winners_2023()
    for uid, st in geo_static["wkr"].items():
        st["erst_winner"] = erst_winners.get(uid) or _leader(_as_frac(st["truth"]))
        # Zweit-Führer nur noch als Meta (kein Proxy mehr)
        z_truth = aggregate_shares(
            precincts,
            {a: precincts[a]["shares"] for a in scope_addrs["wkr"][uid]},
            scope_addrs["wkr"][uid],
        )
        st["zweit_winner"] = _leader(z_truth)
        st["ballot"] = "erst"

    order = [
        a
        for a in simulate_order(
            precincts, scenario, rng, reporting_times=reporting_times
        )
        if a in precincts
    ]
    neighbor_index = _build_neighbor_index(addrs, precincts, struktur, cfg)
    mc_rng = np.random.default_rng(seed + 1)
    scenario_defs = load_be_scenario_defs()
    wkr_bez = {u["id"]: u["bezirk"] for u in geo_units["wkr"]}

    total_g = sum(precincts[a]["gueltig"] for a in addrs)
    steps = []
    targets = np.linspace(0.0, 1.0, n_steps)
    cursor = 0
    reported: set[str] = set()
    reported_g = 0.0
    # Running-min envelopes so ± never widens over the night
    prev_unc_land: dict[str, float] | None = None
    prev_unc_bez: dict[str, dict[str, float] | None] = {}
    prev_unc_wkr: dict[str, dict[str, float] | None] = {}
    prev_unc_turnout: float | None = None

    for tfrac in targets:
        while cursor < len(order) and (reported_g / (total_g + EPS)) < tfrac - 1e-12:
            a = order[cursor]
            reported.add(a)
            reported_g += precincts[a]["gueltig"]
            cursor += 1
        if tfrac >= 1.0 - 1e-12:
            reported = set(addrs)
            reported_g = total_g

        nc, diag = nowcast_at(
            reported,
            precincts,
            prior_shares,
            l1_shares,
            struktur,
            cfg,
            neighbor_index=neighbor_index,
            method="full",
            prior_unc_pp=prior_unc_pp,
        )
        diag["uncertainty"] = monotone_uncertainty(
            prev_unc_land, diag.get("uncertainty")
        )
        prev_unc_land = diag["uncertainty"]
        nc_land = aggregate_shares(precincts, nc)
        if reported:
            naive_land = aggregate_shares(
                precincts, {a: precincts[a]["shares"] for a in reported}, list(reported)
            )
        else:
            naive_land = dict(prior_land)

        by_bezirk = {
            uid: _region_step(
                ua,
                precincts,
                reported,
                nc,
                prior_shares,
                geo_static["bezirk"][uid],
                land_diag=diag,
                prior_unc_pp=prior_unc_pp,
            )
            for uid, ua in sorted(scope_addrs["bezirk"].items())
        }
        # Direktmandate: Erststimmen-Nowcast (nicht Zweit-Proxy)
        nc_erst, diag_erst = nowcast_at(
            reported,
            precincts_erst,
            erst_prior,
            erst_l1_shares,
            struktur,
            cfg,
            neighbor_index=neighbor_index,
            method="full",
            prior_unc_pp=prior_unc_pp,
        )
        by_wkr = {
            uid: _region_step(
                ua,
                precincts_erst,
                reported,
                nc_erst,
                erst_prior,
                geo_static["wkr"][uid],
                land_diag=diag_erst,
                prior_unc_pp=prior_unc_pp,
            )
            for uid, ua in sorted(
                scope_addrs["wkr"].items(), key=lambda kv: int(kv[0])
            )
        }
        for uid, region in by_bezirk.items():
            lp = _leader(_as_frac(region["nowcast"]))
            lt = _leader(_as_frac(geo_static["bezirk"][uid]["truth"]))
            region["leader_pred"] = lp
            region["leader_truth"] = lt
            region["leader_ok"] = lp == lt
            region["uncertainty"] = monotone_uncertainty(
                prev_unc_bez.get(uid), region.get("uncertainty")
            )
            prev_unc_bez[uid] = region["uncertainty"]
        for uid, region in by_wkr.items():
            lp = _leader(_as_frac(region["nowcast"]))
            lt = geo_static["wkr"][uid].get("erst_winner") or _leader(
                _as_frac(geo_static["wkr"][uid]["truth"])
            )
            region["leader_pred"] = lp
            region["leader_truth"] = lt
            region["leader_ok"] = lp == lt
            region["ballot"] = "erst"
            # P(Führung hält): Normal-Approx auf Top-1 vs. Top-2 Marge.
            # Unsicherheit + Prior-Floor solange der WK lokal kaum ausgezählt ist.
            sh = region["nowcast"]
            u = inflate_unc_open_prior(
                region.get("uncertainty"), region["frac_reported"]
            )
            u = monotone_uncertainty(prev_unc_wkr.get(uid), u)
            region["uncertainty"] = u
            prev_unc_wkr[uid] = u
            top2 = sorted(
                MAIN_PARTIES, key=lambda p: sh.get(p, 0.0), reverse=True
            )[:2]
            margin = sh.get(top2[0], 0.0) - sh.get(top2[1], 0.0)
            complete = region["frac_reported"] >= 0.999
            p_lead = p_lead_from_margin(margin, u, top2, complete=complete)
            open_w = max(0.0, 1.0 - float(region["frac_reported"]))
            # Residual veto: narrow lead vs. still-open share (WK69 at 96%).
            residual_safe = complete or (
                open_w <= 0 or margin >= open_w * CALL_RESIDUAL_SWING_PP
            )
            region["runner_up"] = top2[1]
            region["margin"] = round(margin, 2)
            region["p_lead"] = round(p_lead, 4)
            region["complete"] = complete
            region["likely"] = bool(p_lead >= CALL_THRESHOLD)
            region["called"] = bool(
                p_lead >= HARD_CALL_THRESHOLD
                and float(region["frac_reported"]) > 0
                and residual_safe
            )
            # Erst-Nowcast = Direktmandat-Vorhersage
            region["direct_pred"] = lp

        step_eval = build_step_eval(
            nc_land,
            naive_land,
            prior_land,
            truth_land,
            by_bezirk,
            by_wkr,
            geo_static,
            erst_winners_truth=erst_winners,
        )

        entry_mc = night_entry_mc(
            _pct_shares(nc_land),
            diag.get("uncertainty") or {},
            {uid: r["nowcast"] for uid, r in by_bezirk.items()},
            {uid: r["direct_pred"] for uid, r in by_wkr.items()},
            wkr_bez,
            mc_rng,
        )
        scen_probs = night_scenario_probs(
            _pct_shares(nc_land),
            diag.get("uncertainty") or {},
            _pct_shares(truth_land),
            _pct_shares(prior_land),
            mc_rng,
            scenario_defs=scenario_defs,
        )
        turnout = nowcast_turnout(reported, precincts, l1)
        if turnout.get("uncertainty") is not None and prev_unc_turnout is not None:
            turnout["uncertainty"] = round(
                min(float(turnout["uncertainty"]), float(prev_unc_turnout)), 2
            )
        if turnout.get("uncertainty") is not None:
            prev_unc_turnout = float(turnout["uncertainty"])

        frac_now = round(reported_g / (total_g + EPS), 4)
        clock, clock_src = clock_for_step(
            order=order,
            n_reported=len(reported),
            frac=frac_now,
            reporting_times=reporting_times,
            scenario=scenario,
        )

        steps.append(
            {
                "frac_reported": frac_now,
                "n_reported": len(reported),
                "n_total": len(addrs),
                "clock": clock,
                "clock_source": clock_src,
                "nowcast": _pct_shares(nc_land),
                "naive": _pct_shares(naive_land),
                "prior": _pct_shares(prior_land),
                "baseline": _pct_shares(prior_land),
                "truth": _pct_shares(truth_land),
                "mae_nowcast": round(mae(nc_land, truth_land) * 100, 4),
                "mae_naive": round(mae(naive_land, truth_land) * 100, 4),
                "mae_prior": round(mae(prior_land, truth_land) * 100, 4),
                "mae_baseline": round(mae(prior_land, truth_land) * 100, 4),
                "learn_weight": diag["learn_weight"],
                "representativeness": diag["representativeness"],
                "surprise": diag["surprise"],
                "uncertainty": diag["uncertainty"],
                "turnout": turnout,
                "eval": step_eval,
                "entry_mc": entry_mc,
                "scenario_probs": scen_probs,
                "by_bezirk": by_bezirk,
                "by_wkr": by_wkr,
            }
        )

    attach_scenario_p_start(steps)

    # Call-Zusammenfassung je WK: wahrscheinlich / gecallt / voll ausgezählt
    wkr_calls: dict[str, dict] = {}
    for uid in sorted(scope_addrs["wkr"], key=int):
        likely_at = None
        likely_at_clock = None
        likely_at_clock_source = None
        likely_leader = None
        called_at = None
        called_at_clock = None
        called_at_clock_source = None
        complete_at = None
        complete_at_clock = None
        complete_at_clock_source = None
        call_leader = None
        flip_after_call = False
        ua = scope_addrs["wkr"][uid]
        for s in steps:
            r = s["by_wkr"][uid]
            lead = r.get("direct_pred") or r["leader_pred"]
            if likely_at is None and r.get("likely"):
                likely_at = s["frac_reported"]
                likely_at_clock = s.get("clock")
                likely_at_clock_source = s.get("clock_source")
                likely_leader = lead
            if called_at is None and r.get("called"):
                called_at = s["frac_reported"]
                called_at_clock = s.get("clock")
                called_at_clock_source = s.get("clock_source")
                call_leader = lead
            elif called_at is not None and lead != call_leader:
                flip_after_call = True
            if complete_at is None and (
                r.get("complete") or float(r.get("frac_reported", 0.0)) >= 0.999
            ):
                complete_at = s["frac_reported"]
                complete_at_clock, complete_at_clock_source = clock_when_addrs_complete(
                    addrs=ua,
                    order=order,
                    reporting_times=reporting_times,
                    scenario=scenario,
                    land_frac=float(s["frac_reported"]),
                    step_clock=s.get("clock"),
                    step_clock_source=s.get("clock_source"),
                )
        truth_w = erst_winners.get(uid)
        leader_for_eval = call_leader if call_leader is not None else likely_leader
        wkr_calls[uid] = {
            "likely_at": likely_at,
            "likely_at_clock": likely_at_clock,
            "likely_at_clock_source": likely_at_clock_source,
            "likely_leader": likely_leader,
            "called_at": called_at,
            "called_at_clock": called_at_clock,
            "called_at_clock_source": called_at_clock_source,
            "complete_at": complete_at,
            "complete_at_clock": complete_at_clock,
            "complete_at_clock_source": complete_at_clock_source,
            "call_leader": call_leader,
            "flip_after_call": flip_after_call,
            "truth_erst": truth_w,
            "call_correct": (
                (leader_for_eval == truth_w) if leader_for_eval else None
            ),
        }

    meta = prior_meta or {}
    return {
        "election": "AGH2023",
        "baseline": (
            "Vorhersage = proportionaler Swing von AGH2016-Wahlbezirken auf landesweites π₀ (Lead ≈ 0)"
        ),
        "prior_source": meta.get("method"),
        "prior_note": meta.get("note"),
        "scenario": scenario,
        "parties": list(PARTIES),
        "party_labels": PARTY_LABELS,
        "nowcast_methods": [
            {"id": m, "label": NOWCAST_METHOD_LABELS[m]} for m in NOWCAST_METHODS
        ],
        "n_precincts": len(addrs),
        "geo_units": geo_units,
        "geo_static": geo_static,
        "precincts": _precinct_index(precincts, precincts_erst),
        "reporting_order": order,
        "prior_target": {p: round(prior_target[p] * 100, 3) for p in PARTIES},
        "prior_uncertainty_pp": prior_unc_pp,
        "steps": steps,
        "wkr_calls": wkr_calls,
        "call_threshold": CALL_THRESHOLD,
        "hard_call_threshold": HARD_CALL_THRESHOLD,
        "erst_l1_meta": erst_l1_meta,
        "final_mae_nowcast": steps[-1]["mae_nowcast"] if steps else None,
        "generated_at": datetime.now(TZ_BERLIN).strftime("%Y-%m-%d %H:%M:%S"),
        # Official L1 (AGH 2016) vs. this replay of AGH 2023. Live 2026 → AGH 2023.
        "last_election": {
            "year": 2016,
            "label": "AGH 2016",
            "turnout": 66.9,
            "parliament_size": 160,
        },
    }


def run_multi_scenario(n_steps: int = 40) -> dict:
    print("Loading AGH 2023 truth …")
    truth = load_truth_2023()
    print(f"  {len(truth)} precincts")
    print("Loading 2016 L1 on 2023 …")
    l1 = load_baseline_2016_on_2023()
    print(f"  {len(l1)} precincts")
    print("Loading Strukturdaten 2023 …")
    struktur = load_struktur_2023()
    print(f"  {len(struktur)} with features")
    if not PRIOR_2023.exists():
        raise SystemExit(
            f"Missing {PRIOR_2023} — run: python3 code/wahlabend_prior_forecast.py"
        )
    prior_target, prior_meta = load_prior_target(PRIOR_2023)
    prior_unc0 = prior_uncertainty_pp(prior_meta)
    print(
        f"Prior π₀ from {PRIOR_2023.name}: { {p: round(prior_target[p]*100,1) for p in PARTIES} }"
    )
    print(f"Prior ± (pp): {prior_unc0}")
    reporting_times = load_reporting_times_2023()
    print(f"Reporting times from AfS _W_: {len(reporting_times)} precincts")

    scenarios = [
        "actual_times",
        "urne_first",
        "random",
        "small_first",
        "green_first",
        "cdu_first",
    ]
    results = {}
    for sc in scenarios:
        print(f"Replay scenario={sc}")
        results[sc] = run_replay(
            truth,
            l1,
            struktur,
            prior_target,
            prior_meta,
            reporting_times=reporting_times,
            scenario=sc,
            n_steps=n_steps,
        )

    primary = results["actual_times"]
    geo_static = primary.get("geo_static")
    precincts_idx = primary.get("precincts")
    full_steps = {sc: results[sc]["steps"] for sc in scenarios}
    orders = {sc: results[sc]["reporting_order"] for sc in scenarios}

    def land_only(steps: list[dict]) -> list[dict]:
        return [
            {k: v for k, v in s.items() if k not in ("by_bezirk", "by_wkr")}
            for s in steps
        ]

    primary["steps"] = land_only(full_steps["actual_times"])
    primary["precincts"] = precincts_idx
    primary["scenarios"] = {
        sc: {
            "label": sc,
            "steps": full_steps[sc],
            "reporting_order": orders[sc],
            "mae_curve": [s["mae_nowcast"] for s in full_steps[sc]],
            "mae_naive_curve": [s["mae_naive"] for s in full_steps[sc]],
            "wk_hit_curve": [
                s["eval"]["wkr"]["nowcast"]["hit_rate"] for s in full_steps[sc]
            ],
            "wk_hit_naive_curve": [
                s["eval"]["wkr"]["naive"]["hit_rate"] for s in full_steps[sc]
            ],
            "hurdle_acc_curve": [
                s["eval"]["list"]["nowcast"]["hurdle_acc"] for s in full_steps[sc]
            ],
            "frac": [s["frac_reported"] for s in full_steps[sc]],
            "learn_weight": [s["learn_weight"] for s in full_steps[sc]],
            "representativeness": [s["representativeness"] for s in full_steps[sc]],
            "wkr_calls": results[sc]["wkr_calls"],
            "final_mae_nowcast": results[sc]["final_mae_nowcast"],
            "note": (
                "AfS-Datenexport _W_ Datum/Zeit (eingefrorene Enddatei; oft "
                "Nachbearbeitung nach der Wahl, nicht Wahlnacht-Eingang)"
                if sc == "actual_times"
                else None
            ),
        }
        for sc in scenarios
    }
    primary["geo_static"] = geo_static
    primary["direkt_candidates_2026"] = load_direkt_roster_placeholders()
    primary["direkt_candidates_note"] = (
        "Platzhalter statt Personennamen (AGH2023-Replay; 2023-Namen fehlen, "
        "2026-Bewerber wären anachronistisch). Wahrheit = Erststimme-Sieger 2023."
    )
    primary["listen_roster_2026"] = load_listen_roster_placeholders()
    primary["listen_roster_note"] = (
        "Listenstruktur (Positionen / Bezirk vs. Land) aus dem 2026-Schema, "
        "Anzeige nur als Platzhalter. Einzug = Nowcast-Sitze (AGH2023-Anteile), "
        "Wackelbereich = MC-Quantile p10–p90 über das Unsicherheitsband."
    )
    primary["features"] = {"bezirkslisten": True, "listen_einzug": True}
    primary["listen_mode"] = "berlin_mixed"
    # top-level order = default scenario (for convenience)
    primary["reporting_order"] = orders["actual_times"]
    primary["reporting_times_note"] = (
        "actual_times = Sortierung nach Datum/Zeit in Datenexport_AGH2023_Zweitstimme_W_BE.csv. "
        "Im veröffentlichten Endstand wirken die Zeiten oft wie Nachbearbeitung ab 13.02., "
        "nicht wie Meldezeiten in der Wahlnacht."
    )
    primary["scenario_probs_note"] = (
        "Politische Szenarien wie auf der Startseite (Mehrheit, stärkste Kraft, "
        "Hürde). Anzeige: aktuelle Nowcast-P (MC über Unsicherheitsband) und "
        "Wkt vor Auszählungsbeginn (Step 0). Call = P≥50 %. Richtig, wenn Call "
        "der Wahrheit entspricht."
    )
    primary["model"] = {
        "name": "prior_swing_surprise_v1_erst",
        "description": (
            "Zweit (Liste): proportionaler Swing 2016→π₀, Update aus gemeldeten WBs "
            "(lokal + Struktur-Nachbarn). "
            "Erst (Direkt): Prior = OLS resp_E~Z+e_l1 (district_model_coefs) auf dem "
            "Zweit-Prior; e_l1 = 2016-Erst wo Adresse matcht, sonst Zweit-L1; "
            "Update = dieselbe Surprise-Korrektur auf Erststimmen der gemeldeten WBs. "
            "Unsicherheit (±) startet mit Landesprognose-CI, schrumpft mit offenem Anteil."
        ),
        "method": "full",
        "erst_l1_meta": primary.get("erst_l1_meta"),
        "config": {
            "k_neighbors": 40,
            "shrink_half_life_votes": 0.10,
            "mix_local": 0.55,
            "prior_file": str(PRIOR_2023.relative_to(REPO)),
        },
    }
    return primary


def write_panel_csv(path: Path) -> None:
    """Compact panel for debugging: adresse, art, shares truth/base."""
    truth = load_truth_2023()
    base = load_baseline_2016_on_2023()
    addrs = sorted(set(truth) & set(base))
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        header = [
            "adresse",
            "bezirk",
            "art",
            "wkr",
            "gueltig",
            "wber",
        ]
        for p in PARTIES:
            header += [f"truth_{p}", f"base_{p}"]
        w.writerow(header)
        for a in addrs:
            row = [
                a,
                truth[a]["bezirk"],
                truth[a]["art"],
                truth[a]["wkr"],
                truth[a]["gueltig"],
                truth[a]["wber"],
            ]
            for p in PARTIES:
                row += [
                    round(truth[a]["shares"][p] * 100, 3),
                    round(base[a]["shares"][p] * 100, 3),
                ]
            w.writerow(row)
    print(f"Wrote {path} ({len(addrs)} rows)")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--steps", type=int, default=40)
    ap.add_argument("--out", type=Path, default=OUT)
    ap.add_argument(
        "--panel",
        type=Path,
        default=PROCESSED / "panel_agh2023_baseline2016.csv",
    )
    args = ap.parse_args()

    if not (RAW / "DL_BE_AGHBVV2023.xlsx").exists():
        raise SystemExit(f"Missing raw data in {RAW} — run fetch first")

    write_panel_csv(args.panel)
    payload = run_multi_scenario(n_steps=args.steps)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    print(f"Wrote {args.out}")
    # quick summary
    for sc, meta in payload["scenarios"].items():
        mid_i = len(meta["mae_curve"]) // 4  # ~25% counted
        mae_n = meta["mae_curve"][mid_i]
        mae_v = meta["mae_naive_curve"][mid_i]
        print(f"  {sc:12} MAE@~25%: nowcast={mae_n:.2f}  naive={mae_v:.2f}")


if __name__ == "__main__":
    main()
